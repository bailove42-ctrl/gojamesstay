from flask import Flask, render_template, request, redirect, session, flash, url_for, jsonify, send_file
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
from dotenv import load_dotenv
import sqlite3
import re
from datetime import datetime, date, timedelta
import os
from io import BytesIO
import json
import uuid
import qrcode
import smtplib
import html
from email.message import EmailMessage
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

load_dotenv()

app = Flask(__name__)
# ใช้ SECRET_KEY จาก Render/Environment เพื่อให้ลิงก์รีเซ็ตรหัสผ่านปลอดภัยและไม่เปลี่ยนหลัง Deploy
app.secret_key = os.environ.get("SECRET_KEY", "gojames-secret-key")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "database.db")
STATIC_DIR = os.path.join(BASE_DIR, "static")
UPLOAD_DIR = os.path.join(STATIC_DIR, "slips")

os.makedirs(STATIC_DIR, exist_ok=True)
os.makedirs(UPLOAD_DIR, exist_ok=True)

ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "webp"}

# ตั้งค่าอีเมลก่อนใช้งานจริง / บน Render ให้ใส่ใน Environment Variables
# EMAIL_PASSWORD ต้องใช้ App Password ของ Gmail (ไม่ใช่รหัสผ่าน Gmail ปกติ)
EMAIL_SENDER = os.environ.get("EMAIL_SENDER", "").strip()
# Gmail App Password บางครั้งคัดลอกมาเป็นกลุ่มมีเว้นวรรค เช่น "abcd efgh ijkl mnop"
# จึงตัดช่องว่างออกอัตโนมัติ เพื่อให้ใช้ได้ทั้งบน Render และตอนรันในเครื่องจากไฟล์ .env
EMAIL_PASSWORD = os.environ.get("EMAIL_PASSWORD", "").replace(" ", "").strip()
ADMIN_EMAIL = (os.environ.get("ADMIN_EMAIL", "").strip() or EMAIL_SENDER)
SMTP_SERVER = os.environ.get("SMTP_SERVER", "smtp.gmail.com").strip()
try:
    SMTP_PORT = int(os.environ.get("SMTP_PORT", "465"))
except ValueError:
    SMTP_PORT = 465
SMTP_USE_SSL = os.environ.get("SMTP_USE_SSL", "1").strip().lower() not in {"0", "false", "no"}
BASE_URL = (os.environ.get("APP_BASE_URL") or os.environ.get("BASE_URL") or "").rstrip("/")
EMAIL_LOGO_FILENAME = os.environ.get("EMAIL_LOGO_FILENAME", "email_gojames_logo.png").strip()
EMAIL_LOGO_URL = os.environ.get("EMAIL_LOGO_URL", "").strip()

serializer = URLSafeTimedSerializer(app.secret_key)

# PromptPay สำหรับรับเงินจริง
# ถ้าเอาขึ้น GitHub/Render ภายหลัง แนะนำให้ย้ายไปใช้ Environment Variable แทน
PROMPTPAY_ID = os.environ.get("PROMPTPAY_ID", "0953087210")
BANK_ACCOUNT_BANK = "กสิกรไทย"
BANK_ACCOUNT_NUMBER = "0518084623"
BANK_ACCOUNT_NAME = "ธเนศ สมโพยม"

BANKS = {
    "promptpay": {"name": "PromptPay", "short": "PP", "color": "#1f8bff", "prefixes": []},
    "kbank": {"name": "กสิกรไทย", "short": "KB", "color": "#0f9d58", "prefixes": ["004"]},
    "scb": {"name": "ไทยพาณิชย์", "short": "SCB", "color": "#6f42c1", "prefixes": ["014"]},
    "bbl": {"name": "กรุงเทพ", "short": "BBL", "color": "#1d4ed8", "prefixes": ["002"]},
    "ktb": {"name": "กรุงไทย", "short": "KTB", "color": "#00a2ff", "prefixes": ["006"]},
    "ttb": {"name": "ttb", "short": "TTB", "color": "#f97316", "prefixes": ["011"]},
    "bay": {"name": "กรุงศรีอยุธยา", "short": "BAY", "color": "#facc15", "prefixes": ["025"]},
    "gsb": {"name": "ออมสิน", "short": "GSB", "color": "#ec4899", "prefixes": ["030"]},
    "baac": {"name": "ธ.ก.ส.", "short": "BAAC", "color": "#22c55e", "prefixes": ["034"]},
    "cimb": {"name": "CIMB Thai", "short": "CIMB", "color": "#ef4444", "prefixes": ["018"]},
    "lhbank": {"name": "LH Bank", "short": "LH", "color": "#6b7280", "prefixes": ["073"]},
    "uob": {"name": "UOB", "short": "UOB", "color": "#1e3a8a", "prefixes": ["024"]},
    "icbc": {"name": "ICBC Thai", "short": "ICBC", "color": "#dc2626", "prefixes": ["070"]},
    "kkp": {"name": "เกียรตินาคินภัทร", "short": "KKP", "color": "#0f766e", "prefixes": ["069"]},
}



def _pp_tlv(tag, value):
    value = str(value)
    return f"{tag}{len(value):02d}{value}"


def _pp_crc16(payload):
    crc = 0xFFFF
    for ch in payload.encode("utf-8"):
        crc ^= ch << 8
        for _ in range(8):
            if crc & 0x8000:
                crc = ((crc << 1) ^ 0x1021) & 0xFFFF
            else:
                crc = (crc << 1) & 0xFFFF
    return f"{crc:04X}"


def _format_promptpay_target(target):
    digits = "".join(ch for ch in str(target) if ch.isdigit())
    if len(digits) == 10 and digits.startswith("0"):   # เบอร์มือถือไทย
        return "0066" + digits[1:]
    if len(digits) == 13:  # เลขบัตรประชาชน
        return digits
    raise ValueError("รูปแบบ PromptPay ID ไม่ถูกต้อง")


def generate_promptpay_payload(target, amount):
    formatted_target = _format_promptpay_target(target)

    # 01 = เบอร์โทรศัพท์, 02 = เลขบัตรประชาชน
    target_tag = "01" if formatted_target.startswith("0066") else "02"

    merchant_account_info = (
        _pp_tlv("00", "A000000677010111") +
        _pp_tlv(target_tag, formatted_target)
    )

    payload = (
        _pp_tlv("00", "01") +          # Payload Format Indicator
        _pp_tlv("01", "12") +          # Point of Initiation Method (Dynamic QR)
        _pp_tlv("29", merchant_account_info) +
        _pp_tlv("52", "0000") +        # Merchant Category Code
        _pp_tlv("53", "764") +         # Currency: THB
        _pp_tlv("54", f"{float(amount):.2f}") +
        _pp_tlv("58", "TH") +
        "6304"
    )
    return payload + _pp_crc16(payload)

ROOMS = [
    {
        "id": 1,
        "name": "ห้อง Lavender",
        "price": 800,
        "detail": "ห้อง Lavender เป็นห้องพักโทนอบอุ่นที่ให้ความรู้สึกสบายตั้งแต่ก้าวแรก ภายในตกแต่งเรียบง่าย โปร่ง และเหมาะกับการพักผ่อนแบบเป็นส่วนตัว มีเตียงใหญ่ เครื่องปรับอากาศ ตู้เย็น Wi‑Fi ฟรี และห้องน้ำส่วนตัวครบถ้วน ด้านหน้าห้องมีมุมพักผ่อนให้นั่งเล่น รับลม หรือจิบเครื่องดื่มเบา ๆ ท่ามกลางบรรยากาศธรรมชาติ เหมาะสำหรับผู้เข้าพัก 2 ท่านที่ต้องการห้องพักเงียบ ๆ ดูแลตัวเองได้สะดวก และเดินทางมาพักผ่อนได้แบบสบายใจ",
        "cover": "room1_1.jpg",
        "images": ["room1_1.jpg", "room1_2.jpg", "room1_3.jpg"],
        "amenities": [
            {"icon": "wifi", "label": "Wi-Fi ฟรี"},
            {"icon": "air", "label": "เครื่องปรับอากาศ"},
            {"icon": "fridge", "label": "ตู้เย็น"},
            {"icon": "bed", "label": "เตียงใหญ่ 1 เตียง"},
            {"icon": "bath", "label": "ห้องน้ำส่วนตัว"}
        ]
    },
    {
        "id": 2,
        "name": "ห้อง Lilac",
        "price": 800,
        "detail": "ห้อง Lilac เหมาะสำหรับคนที่ชอบบรรยากาศสวนและความเป็นธรรมชาติ รอบห้องให้ความรู้สึกสงบ มีพื้นที่นั่งเล่นหน้าห้องสำหรับพักสายตา พูดคุย หรือใช้เวลาช้า ๆ ในวันหยุด ภายในห้องจัดวางอย่างเรียบง่ายแต่ใช้งานสะดวก มี Wi‑Fi ฟรี เครื่องปรับอากาศ ตู้เย็น เตียงใหญ่ 1 เตียง และห้องน้ำส่วนตัว เหมาะกับคู่รักหรือผู้เข้าพัก 2 ท่านที่อยากได้ห้องพักสบาย ๆ มีมุมส่วนตัว และอยู่ใกล้บรรยากาศร่มรื่นของสวน",
        "cover": "room2_1.jpg",
        "images": ["room2_1.jpg", "room2_2.jpg", "room2_3.jpg"],
        "amenities": [
            {"icon": "wifi", "label": "Wi-Fi ฟรี"},
            {"icon": "air", "label": "เครื่องปรับอากาศ"},
            {"icon": "fridge", "label": "ตู้เย็น"},
            {"icon": "bed", "label": "เตียงใหญ่ 1 เตียง"},
            {"icon": "bath", "label": "ห้องน้ำส่วนตัว"}
        ]
    },
    {
        "id": 3,
        "name": "ห้อง Orchid",
        "price": 800,
        "detail": "ห้อง Orchid เป็นห้องพักสไตล์บ้านพักส่วนตัวที่ให้บรรยากาศเรียบง่ายและผ่อนคลาย รายล้อมด้วยความร่มรื่น เหมาะสำหรับคนที่ต้องการหลีกหนีความวุ่นวายและพักผ่อนอย่างเป็นกันเอง ภายในมีพื้นที่ใช้งานครบสำหรับการเข้าพัก 2 ท่าน พร้อม Wi‑Fi ฟรี เครื่องปรับอากาศ ตู้เย็น เตียงใหญ่ 1 เตียง และห้องน้ำส่วนตัว บริเวณหน้าห้องให้ความรู้สึกโปร่ง สบาย และมีความเป็นส่วนตัว เหมาะกับการมาพักระยะสั้นหรือพักผ่อนในวันหยุดแบบเงียบสงบ",
        "cover": "room3_1.jpg",
        "images": ["room3_1.jpg", "room3_2.jpg", "room3_3.jpg"],
        "amenities": [
            {"icon": "wifi", "label": "Wi-Fi ฟรี"},
            {"icon": "air", "label": "เครื่องปรับอากาศ"},
            {"icon": "fridge", "label": "ตู้เย็น"},
            {"icon": "bed", "label": "เตียงใหญ่ 1 เตียง"},
            {"icon": "bath", "label": "ห้องน้ำส่วนตัว"}
        ]
    },
    {
        "id": 4,
        "name": "ห้อง Iris",
        "price": 800,
        "detail": "ห้องพักสำหรับครอบครัวขนาดเล็ก ดูเรียบง่ายสบายตา",
        "cover": "room4_real.jpg",
        "images": ["room4_1.svg", "room4_2.svg", "room4_3.svg"],
        "is_open": False,
        "status_text": "ยังไม่พร้อมให้บริการ",
        "maintenance_note": "ห้องนี้ยังไม่พร้อมให้บริการ และจะเปิดให้จองในภายหลัง"
    },
    {
        "id": 5,
        "name": "ห้อง Violet",
        "price": 800,
        "detail": "ห้องพรีเมียม พร้อมระเบียงและพื้นที่กว้างขึ้น",
        "cover": "room5_real.jpg",
        "images": ["room5_1.svg", "room5_2.svg", "room5_3.svg"],
        "is_open": False,
        "status_text": "ยังไม่พร้อมให้บริการ",
        "maintenance_note": "ห้องนี้ยังไม่พร้อมให้บริการ และจะเปิดให้จองในภายหลัง"
    }
]

BOOKED_STATUSES = (
    "approved",
    "waiting_slip_approval",
    "deposit_paid",
    "waiting_remaining_slip_approval",
    "remaining_counter_pending",
    "counter_pending",
    "fully_paid",
    "checked_in",
    "cancel_requested",
)
NO_SHOW_CUTOFF_TIME = "20:00"
RESORT_CHECKIN_GUIDE = "14:00"
RESORT_CHECKOUT_GUIDE = "12:00"  # หลังเวลานี้ หากลูกค้ายังไม่มา ผู้ดูแลระบบสามารถกด No Show เพื่อปล่อยห้องได้

STANDARD_CHECKIN_TIME = "14:00"
STANDARD_CHECKOUT_TIME = "12:00"
HOTEL_CHECKIN_TIME_OPTIONS = [f"{hour:02d}:00" for hour in range(14, 23)]  # 14:00-22:00
HOTEL_CHECKOUT_TIME_OPTIONS = [f"{hour:02d}:00" for hour in range(6, 13)]  # 06:00-12:00


MANUAL_DAY_STATUSES = {
    "free": "ว่าง",
    "booked": "มีผู้จอง",
    "occupied": "มีผู้พักอยู่",
    "blocked": "ไม่ว่าง"
}


def make_receipt_no(booking_id):
    return f"RCPT-{datetime.now().strftime('%Y%m%d')}-{booking_id:05d}"


def ensure_walkin_user(phone, display_name):
    """สร้าง/ใช้ user สำหรับ Walk-in เพื่อให้ระบบจองยังอ้างอิง username ได้"""
    username = "walkin_" + re.sub(r"[^0-9A-Za-z_]+", "", phone or "guest")
    if username == "walkin_":
        username = "walkin_guest"

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT username FROM users WHERE username = ?", (username,))
    exists = cur.fetchone()
    if not exists:
        cur.execute("""
            INSERT INTO users(first_name,last_name,nickname,gender,phone,email,username,password,role)
            VALUES(?,?,?,?,?,?,?,?,?)
        """, (
            display_name or "Walk-in",
            "",
            display_name or "Walk-in",
            "-",
            phone or "-",
            f"{username}@walkin.local",
            username,
            "walkin",
            "user"
        ))
        conn.commit()
    conn.close()
    return username


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def _safe_json_list(value, fallback=None):
    if fallback is None:
        fallback = []
    if not value:
        return fallback
    try:
        data = json.loads(value)
        return data if isinstance(data, list) else fallback
    except Exception:
        return fallback


def room_row_to_dict(row):
    room = dict(row)
    room["is_open"] = bool(room.get("is_open", 1))
    room["is_deleted"] = bool(room.get("is_deleted", 0))
    room["price"] = float(room.get("price") or 0)
    room["images"] = _safe_json_list(room.get("images"), [room.get("cover") or "room1_1.jpg"])
    room["amenities"] = _safe_json_list(room.get("amenities"), [
        {"icon": "wifi", "label": "Wi-Fi ฟรี"},
        {"icon": "air", "label": "เครื่องปรับอากาศ"},
        {"icon": "bath", "label": "ห้องน้ำส่วนตัว"}
    ])
    room["amenity_labels"] = [item.get("label", "") if isinstance(item, dict) else str(item) for item in room.get("amenities", [])]
    return room


def get_room_status_map():
    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute("SELECT room_id, is_open FROM room_settings")
        rows = cur.fetchall()
    except sqlite3.OperationalError:
        rows = []
    conn.close()
    return {row["room_id"]: bool(row["is_open"]) for row in rows}


def get_rooms_with_status(include_deleted=False):
    conn = get_db()
    cur = conn.cursor()
    try:
        if include_deleted:
            cur.execute("SELECT * FROM rooms ORDER BY id ASC")
        else:
            cur.execute("SELECT * FROM rooms WHERE COALESCE(is_deleted,0)=0 ORDER BY id ASC")
        rows = cur.fetchall()
        conn.close()
        if rows:
            return [room_row_to_dict(row) for row in rows]
    except sqlite3.OperationalError:
        conn.close()

    # fallback กรณีฐานข้อมูลยังไม่ได้สร้างตาราง rooms
    status_map = get_room_status_map()
    rooms = []
    for room in ROOMS:
        room_copy = dict(room)
        default_open = room_copy.get("is_open", True)
        room_copy["is_open"] = status_map.get(room_copy["id"], default_open)
        room_copy["is_deleted"] = False
        room_copy["amenity_labels"] = [item.get("label", "") if isinstance(item, dict) else str(item) for item in room_copy.get("amenities", [])]
        rooms.append(room_copy)
    return rooms


def save_uploaded_room_file(file_storage):
    if not file_storage or not file_storage.filename:
        return ""
    if not allowed_file(file_storage.filename):
        return ""
    filename = secure_filename(file_storage.filename)
    ext = filename.rsplit(".", 1)[1].lower() if "." in filename else "jpg"
    final_name = f"room_{uuid.uuid4().hex[:12]}.{ext}"
    file_storage.save(os.path.join(STATIC_DIR, final_name))
    return final_name



ROOM_AMENITY_OPTIONS = [
    ("ภายในห้องพัก", [
        "Wi-Fi ฟรี",
        "เครื่องปรับอากาศ",
        "โทรทัศน์",
        "ตู้เย็น",
        "เครื่องทำน้ำอุ่น",
        "ห้องน้ำส่วนตัว",
        "เตียงใหญ่ 1 เตียง",
        "เตียงเดี่ยว 2 เตียง",
        "กาต้มน้ำไฟฟ้า",
        "ไดร์เป่าผม",
        "ผ้าเช็ดตัว",
        "สบู่ / แชมพู",
        "โต๊ะทำงาน",
        "ตู้เสื้อผ้า",
    ]),
    ("พื้นที่ส่วนกลาง", [
        "ที่จอดรถฟรี",
        "กล้องวงจรปิด (CCTV)",
        "สวนพักผ่อน",
        "พื้นที่นั่งเล่นส่วนกลาง",
        "ระเบียงหน้าห้อง",
        "มุมถ่ายรูป",
    ]),
    ("ตัวเลือกเพิ่มเติม", [
        "อาหารเช้า",
        "เตียงเสริมได้",
        "ห้องสำหรับครอบครัว",
        "พักได้หลายคน",
        "นำสัตว์เลี้ยงเข้าได้",
        "สูบบุหรี่ได้",
        "เตาปิ้งย่าง",
        "คาราโอเกะ",
    ]),
]

KNOWN_AMENITY_LABELS = [label for _, options in ROOM_AMENITY_OPTIONS for label in options]

# จับคู่ชื่อสิ่งอำนวยความสะดวกกับไอคอนที่ใช้แสดงบนหน้ารายละเอียดห้อง
# ถ้าเป็นสิ่งอำนวยความสะดวกที่พิมพ์เพิ่มเอง ระบบจะใช้ไอคอน check เป็นค่าเริ่มต้น
AMENITY_ICON_MAP = {
    "Wi-Fi ฟรี": "wifi",
    "เครื่องปรับอากาศ": "air",
    "โทรทัศน์": "tv",
    "ตู้เย็น": "fridge",
    "เครื่องทำน้ำอุ่น": "shower",
    "กาต้มน้ำไฟฟ้า": "kettle",
    "ไดร์เป่าผม": "dryer",
    "ผ้าเช็ดตัว": "towel",
    "สบู่ / แชมพู": "soap",
    "ห้องน้ำส่วนตัว": "bath",
    "เตียงใหญ่ 1 เตียง": "bed",
    "เตียงเดี่ยว 2 เตียง": "bed",
    "เตียงเสริมได้": "bed",
    "ที่จอดรถฟรี": "parking",
    "กล้องวงจรปิด (CCTV)": "camera",
    "สวนพักผ่อน": "tree",
    "พื้นที่นั่งเล่นส่วนกลาง": "sofa",
    "ระเบียงหน้าห้อง": "tree",
    "มุมถ่ายรูป": "camera",
    "อาหารเช้า": "utensils",
    "ห้องสำหรับครอบครัว": "family",
    "นำสัตว์เลี้ยงเข้าได้": "paw",
    "สูบบุหรี่ได้": "smoking",
    "พักได้หลายคน": "family",
    "โต๊ะทำงาน": "check",
    "ตู้เสื้อผ้า": "check",
    "เตาปิ้งย่าง": "utensils",
    "คาราโอเกะ": "check",
}

def amenity_icon_for(label):
    label = (label or "").strip()
    return AMENITY_ICON_MAP.get(label, "check")


def build_room_amenities_from_form():
    selected = []
    seen = set()
    for label in request.form.getlist("amenities_options"):
        label = (label or "").strip()
        if label and label not in seen:
            selected.append({"icon": amenity_icon_for(label), "label": label})
            seen.add(label)

    extra_text = request.form.get("amenities_extra", "")
    for line in (extra_text or "").splitlines():
        label = line.strip(" -•\t")
        if label and label not in seen:
            selected.append({"icon": amenity_icon_for(label), "label": label})
            seen.add(label)

    if not selected:
        return parse_room_amenities("")
    return selected

def parse_room_amenities(text):
    items = []
    for line in (text or "").splitlines():
        label = line.strip(" -•\t")
        if label:
            items.append({"icon": amenity_icon_for(label), "label": label})
    if not items:
        items = [
            {"icon": "wifi", "label": "Wi-Fi ฟรี"},
            {"icon": "air", "label": "เครื่องปรับอากาศ"},
            {"icon": "bath", "label": "ห้องน้ำส่วนตัว"}
        ]
    return items


def get_no_show_cutoff_dt(booking):
    """เวลาตัดสิทธิ์ No Show: ใช้วันที่เข้าพัก + 20:00 น. เป็นค่าเริ่มต้น"""
    try:
        cutoff_time = NO_SHOW_CUTOFF_TIME
        return datetime.strptime(f"{booking['checkin']} {cutoff_time}", "%Y-%m-%d %H:%M")
    except Exception:
        try:
            return datetime.strptime(f"{booking['checkin']} {booking['checkin_time']}", "%Y-%m-%d %H:%M")
        except Exception:
            return None


def get_now():
    return datetime.now()


def db_now_str():
    return get_now().strftime("%Y-%m-%d %H:%M:%S.%f")


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def normalize_digits(value):
    return "".join(ch for ch in str(value or "") if ch.isdigit())


def get_bank_meta(bank_code):
    meta = BANKS.get((bank_code or "").lower())
    if meta:
        return meta
    return {"name": bank_code or "-", "short": (bank_code or "-")[:4].upper(), "color": "#6b7280", "prefixes": []}


def detect_bank_from_account(account_number):
    digits = normalize_digits(account_number)
    if len(digits) in (10, 13) and digits.startswith("0"):
        return "promptpay"
    prefix = digits[:3]
    for code, meta in BANKS.items():
        if code == "promptpay":
            continue
        if prefix in meta.get("prefixes", []):
            return code
    return ""


def validate_refund_destination(refund_method, refund_account, refund_bank=""):
    method = (refund_method or "").lower().strip()
    digits = normalize_digits(refund_account)

    if method == "promptpay":
        if (len(digits) == 10 and digits.startswith("0")) or len(digits) == 13:
            return True, "promptpay", digits, ""
        return False, "", digits, "PromptPay ต้องเป็นเบอร์โทร 10 หลักหรือเลขบัตรประชาชน 13 หลัก"

    if method == "bank":
        if not digits.isdigit() or len(digits) < 10 or len(digits) > 12:
            return False, "", digits, "เลขบัญชีธนาคารต้องเป็นตัวเลข 10-12 หลัก"
        detected_bank = detect_bank_from_account(digits)
        selected_bank = (refund_bank or "").lower().strip()
        final_bank = selected_bank or detected_bank
        if not final_bank:
            return False, "", digits, "ไม่สามารถตรวจจับธนาคารจากเลขบัญชีนี้ได้ กรุณาเลือกธนาคาร"
        if selected_bank and detected_bank and selected_bank != detected_bank:
            return False, "", digits, "เลขบัญชีนี้ดูเหมือนจะไม่ตรงกับธนาคารที่เลือก กรุณาตรวจสอบอีกครั้ง"
        return True, final_bank, digits, ""

    return False, "", digits, "กรุณาเลือกวิธีรับเงินคืน"


app.jinja_env.globals["get_bank_meta"] = get_bank_meta
app.jinja_env.globals["KNOWN_AMENITY_LABELS"] = KNOWN_AMENITY_LABELS


def build_email_html(subject, body, action_url=None, action_label="เปิดเว็บไซต์", logo_url=None):
    """สร้าง HTML Email แบบการ์ดสวย ๆ โดยใช้โลโก้จาก URL แทนการแนบไฟล์

    เหตุผล: ถ้าแนบโลโก้แบบ inline cid บางเมล เช่น Gmail จะแสดงเป็นไฟล์แนบ noname
    การใช้ URL จาก /static/... จะทำให้โลโก้อยู่ในเนื้อหาเมลและไม่ขึ้นเป็นไฟล์แนบ
    """
    body = str(body or "")
    subject_safe = html.escape(str(subject or "แจ้งเตือนจาก GoJames"))

    if logo_url:
        logo_url_safe = html.escape(logo_url, quote=True)
        logo_html = (
            f'<img src="{logo_url_safe}" alt="GoJames" '
            'style="display:block;margin:0 auto 8px auto;width:220px;max-width:72%;height:auto;border:0;outline:none;text-decoration:none;">'
        )
    else:
        logo_html = '<div style="font-size:28px;font-weight:800;color:#ffffff;letter-spacing:.2px;">GoJames</div>'

    url_pattern = re.compile(r"https?://\S+")
    urls = url_pattern.findall(body)
    if not action_url and urls:
        action_url = urls[0].rstrip(".,)>")

    # ซ่อน URL ยาวในส่วนเนื้อหา แล้วไปแสดงเป็นปุ่มแทน
    body_for_lines = url_pattern.sub("", body).strip()
    lines = [line.strip() for line in body_for_lines.splitlines()]
    paragraphs = []
    for line in lines:
        if line:
            paragraphs.append(
                f'<p style="margin:0 0 12px 0;color:#2b1747;font-size:15px;line-height:1.75;">{html.escape(line)}</p>'
            )
        else:
            paragraphs.append('<div style="height:6px;"></div>')

    if not paragraphs:
        paragraphs.append(
            '<p style="margin:0;color:#2b1747;font-size:15px;line-height:1.75;">มีการแจ้งเตือนจากระบบ GoJames Vacation Home</p>'
        )

    button_html = ""
    if action_url:
        action_url_safe = html.escape(action_url, quote=True)
        action_label_safe = html.escape(action_label or "เปิดลิงก์")
        button_html = f"""
        <div style="text-align:center;margin:28px 0 10px 0;">
            <a href="{action_url_safe}" style="background:#7c3aed;color:#ffffff;text-decoration:none;padding:13px 28px;border-radius:12px;font-weight:700;display:inline-block;box-shadow:0 10px 22px rgba(124,58,237,.22);">{action_label_safe}</a>
        </div>
        <p style="margin:14px 0 0 0;color:#7c6a95;font-size:12px;line-height:1.6;word-break:break-all;text-align:center;">หากกดปุ่มไม่ได้ ให้คัดลอกลิงก์นี้ไปเปิดในเบราว์เซอร์:<br>{action_url_safe}</p>
        """

    content_html = "\n".join(paragraphs)
    return f"""<!doctype html>
<html lang="th">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
</head>
<body style="margin:0;padding:0;background:#f4edff;font-family:Tahoma,Arial,sans-serif;">
  <div style="max-width:640px;margin:0 auto;padding:28px 14px;">
    <div style="background:#ffffff;border:1px solid #eadcff;border-radius:22px;overflow:hidden;box-shadow:0 18px 45px rgba(50,24,79,.10);">
      <div style="background:linear-gradient(135deg,#2c183f,#7c3aed);padding:24px 26px;text-align:center;">
        {logo_html}
        <div style="font-size:13px;color:#eadcff;margin-top:4px;font-weight:700;">GoJames Vacation Home</div>
      </div>
      <div style="padding:30px 28px;">
        <div style="display:inline-block;background:#f0e6ff;color:#6d28d9;border-radius:999px;padding:6px 12px;font-size:12px;font-weight:700;margin-bottom:14px;">แจ้งเตือนจากระบบ</div>
        <h2 style="margin:0 0 18px 0;color:#1f0f35;font-size:24px;line-height:1.35;">{subject_safe}</h2>
        {content_html}
        {button_html}
        <div style="margin-top:28px;border-top:1px solid #eadcff;padding-top:18px;">
          <p style="margin:0;color:#8b7aa8;font-size:12px;line-height:1.6;">อีเมลนี้ถูกส่งอัตโนมัติจากระบบ GoJames Vacation Home กรุณาอย่าตอบกลับอีเมลฉบับนี้</p>
        </div>
      </div>
    </div>
  </div>
</body>
</html>"""


def get_email_logo_url():
    """คืน URL โลโก้สำหรับใส่ใน HTML Email โดยไม่แนบเป็นไฟล์"""
    if EMAIL_LOGO_URL:
        return EMAIL_LOGO_URL
    try:
        base_url = BASE_URL or request.url_root.rstrip("/")
    except RuntimeError:
        base_url = BASE_URL
    if base_url and EMAIL_LOGO_FILENAME:
        return f"{base_url}/static/{EMAIL_LOGO_FILENAME}"
    return ""


def send_email_notification(to_email, subject, body, html_body=None, action_url=None, action_label="เปิดเว็บไซต์"):
    """ส่งอีเมลผ่าน SMTP และคืนค่า True/False เพื่อให้หน้าเว็บแจ้งผลได้ถูกต้อง

    ใช้ได้ทั้งบน Render ผ่าน Environment Variables และในเครื่องผ่านไฟล์ .env
    ถ้าส่งไม่สำเร็จ จะพิมพ์ Error ลง Console/Render Logs และไม่ทำให้เว็บค้างนาน
    """
    to_email = (to_email or "").strip()
    if not to_email:
        print("[EMAIL] ไม่ได้ระบุผู้รับ")
        return False

    if not EMAIL_SENDER or not EMAIL_PASSWORD:
        print("[EMAIL] ยังไม่ได้ตั้งค่า EMAIL_SENDER หรือ EMAIL_PASSWORD")
        return False

    try:
        msg = EmailMessage()
        msg["Subject"] = subject
        msg["From"] = EMAIL_SENDER
        msg["To"] = to_email
        msg.set_content(str(body or ""))
        msg.add_alternative(
            html_body if html_body else build_email_html(
                subject,
                body,
                action_url=action_url,
                action_label=action_label,
                logo_url=get_email_logo_url()
            ),
            subtype="html"
        )

        if SMTP_USE_SSL:
            with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT, timeout=8) as smtp:
                smtp.login(EMAIL_SENDER, EMAIL_PASSWORD)
                smtp.send_message(msg)
        else:
            with smtplib.SMTP(SMTP_SERVER, SMTP_PORT, timeout=8) as smtp:
                smtp.ehlo()
                smtp.starttls()
                smtp.ehlo()
                smtp.login(EMAIL_SENDER, EMAIL_PASSWORD)
                smtp.send_message(msg)

        print(f"[EMAIL] ส่งสำเร็จ -> {to_email}: {subject}")
        return True
    except Exception as e:
        print(f"[EMAIL] ส่งอีเมลไม่สำเร็จ -> {to_email}: {type(e).__name__}: {e}")
        return False

def send_reset_password_email(user_row):
    token = serializer.dumps({"user_id": user_row["id"], "email": user_row["email"]}, salt="reset-password")
    base_url = BASE_URL or request.url_root.rstrip("/")
    reset_link = f"{base_url}{url_for('reset_password', token=token)}"

    subject = "รีเซ็ตรหัสผ่าน - GoJames Vacation Home"
    body = (
        f"สวัสดี {user_row['username']}\n\n"
        "มีการขอรีเซ็ตรหัสผ่านสำหรับบัญชี GoJames Vacation Home ของคุณ\n"
        "กดลิงก์ด้านล่างเพื่อตั้งรหัสผ่านใหม่\n\n"
        f"{reset_link}\n\n"
        "ลิงก์นี้จะหมดอายุภายใน 30 นาที\n"
        "หากคุณไม่ได้เป็นผู้ขอรีเซ็ต สามารถละเว้นอีเมลฉบับนี้ได้"
    )
    return send_email_notification(user_row["email"], subject, body, action_url=reset_link, action_label="ตั้งรหัสผ่านใหม่")


def create_notification(username, title, message):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO notifications(username, title, message, is_read, created_at)
        VALUES(?,?,?,?,?)
    """, (username, title, message, 0, db_now_str()))
    conn.commit()
    conn.close()


def notify_user(username, title, message, email_subject=None, email_body=None):
    create_notification(username, title, message)

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT email FROM users WHERE username = ?", (username,))
    user = cur.fetchone()
    conn.close()

    if user and user["email"]:
        send_email_notification(
            user["email"],
            email_subject if email_subject else title,
            email_body if email_body else message
        )


def notify_admin(title, message, email_subject=None, email_body=None):
    """แจ้งเตือนผู้ดูแลระบบทั้งในเว็บและทางอีเมล (ถ้าตั้งค่า ADMIN_EMAIL หรืออีเมลของ admin ไว้)"""
    create_notification("admin", title, message)

    recipients = []
    if ADMIN_EMAIL:
        recipients.append(ADMIN_EMAIL)

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT email FROM users WHERE username = ?", ("admin",))
    admin_user = cur.fetchone()
    conn.close()

    if admin_user and admin_user["email"]:
        recipients.append(admin_user["email"])

    seen = set()
    for email in recipients:
        email = (email or "").strip()
        if not email or email.lower() in seen:
            continue
        seen.add(email.lower())
        send_email_notification(
            email,
            email_subject if email_subject else title,
            email_body if email_body else message
        )



@app.route("/admin/test-email")
def admin_test_email():
    """ทดสอบระบบส่งอีเมลจากหน้าผู้ดูแลระบบ ใช้ได้ทั้ง Local และ Render"""
    if session.get("role") != "admin":
        return redirect("/login")

    target = ADMIN_EMAIL or EMAIL_SENDER
    site_url = BASE_URL or request.url_root.rstrip('/')
    ok = send_email_notification(
        target,
        "ทดสอบส่งอีเมล - GoJames Vacation Home",
        "หากได้รับอีเมลนี้ แสดงว่าระบบส่งอีเมลของ GoJames ใช้งานได้แล้ว\n\n"
        f"เว็บ: {site_url}",
        action_url=site_url,
        action_label="เปิดเว็บไซต์"
    )
    if ok:
        flash(f"ส่งอีเมลทดสอบไปที่ {target} แล้ว")
    else:
        flash("ส่งอีเมลทดสอบไม่สำเร็จ กรุณาตรวจสอบ EMAIL_SENDER / EMAIL_PASSWORD / Render Logs")
    return redirect("/admin")


def status_th(status):
    return {
        "waiting_approval": "รอชำระเงิน",
        "approved": "รอชำระเงิน",
        "waiting_slip_approval": "รอตรวจสอบการชำระเงิน",
        "deposit_paid": "มัดจำแล้ว",
        "waiting_remaining_slip_approval": "รอตรวจสอบการชำระเงินส่วนที่เหลือ",
        "remaining_counter_pending": "รอชำระส่วนที่เหลือหน้าเคาน์เตอร์",
        "counter_pending": "รอชำระหน้าเคาน์เตอร์",
        "fully_paid": "ชำระเงินครบแล้ว",
        "checked_in": "เข้าพักแล้ว",
        "checked_out": "เช็กเอาท์แล้ว",
        "cancel_requested": "ขอยกเลิก",
        "refund_pending": "รอคืนเงิน",
        "refunded": "คืนเงินแล้ว",
        "cancelled": "ยกเลิกแล้ว",
        "rejected": "ถูกปฏิเสธ",
        "no_show": "ลูกค้าไม่มา",
    }.get(status, status or "-")

@app.context_processor
def inject_notification_count():
    unread_count = 0
    latest_notifications = []
    user_nav_badge_count = 0
    admin_nav_badge_count = 0

    if session.get("user"):
        conn = get_db()
        cur = conn.cursor()

        # badge ของ "แจ้งเตือน" ใช้จาก unread ปกติ
        cur.execute("""
            SELECT COUNT(*) AS total FROM notifications
            WHERE username = ? AND is_read = 0
        """, (session["user"],))
        row = cur.fetchone()
        unread_count = row["total"] if row else 0

        cur.execute("""
            SELECT * FROM notifications
            WHERE username = ?
            ORDER BY id DESC
            LIMIT 5
        """, (session["user"],))
        latest_notifications = cur.fetchall()

        # badge ของเมนู "การจองของฉัน" / "จัดการการจอง"
        cur.execute("""
            SELECT last_seen_user_notification_id, last_seen_admin_notification_id
            FROM users
            WHERE username = ?
        """, (session["user"],))
        seen_row = cur.fetchone()

        if session.get("role") == "admin":
            last_seen_admin_id = seen_row["last_seen_admin_notification_id"] if seen_row else 0
            cur.execute("""
                SELECT COUNT(*) AS total FROM notifications
                WHERE username = 'admin' AND id > ?
            """, (last_seen_admin_id or 0,))
            row = cur.fetchone()
            admin_nav_badge_count = row["total"] if row else 0
        else:
            last_seen_user_id = seen_row["last_seen_user_notification_id"] if seen_row else 0
            cur.execute("""
                SELECT COUNT(*) AS total FROM notifications
                WHERE username = ? AND id > ?
            """, (session["user"], last_seen_user_id or 0))
            row = cur.fetchone()
            user_nav_badge_count = row["total"] if row else 0
        conn.close()

    return {
        "unread_notification_count": unread_count,
        "latest_notifications": latest_notifications,
        "user_nav_badge_count": user_nav_badge_count,
        "admin_nav_badge_count": admin_nav_badge_count,
        "status_th": status_th
    }

def get_live_badge_payload():
    unread_count = 0
    user_nav_badge_count = 0
    admin_nav_badge_count = 0

    if session.get("user"):
        conn = get_db()
        cur = conn.cursor()

        cur.execute("""
            SELECT COUNT(*) AS total FROM notifications
            WHERE username = ? AND is_read = 0
        """, (session["user"],))
        row = cur.fetchone()
        unread_count = row["total"] if row else 0

        cur.execute("""
            SELECT last_seen_user_bookings, last_seen_admin_bookings
            FROM users
            WHERE username = ?
        """, (session["user"],))
        seen_row = cur.fetchone()

        if session.get("role") == "admin":
            last_seen_admin = seen_row["last_seen_admin_bookings"] if seen_row else None
            if last_seen_admin:
                cur.execute("""
                    SELECT COUNT(*) AS total FROM notifications
                    WHERE username = 'admin' AND created_at > ?
                """, (last_seen_admin,))
            else:
                cur.execute("""
                    SELECT COUNT(*) AS total FROM notifications
                    WHERE username = 'admin'
                """)
            row = cur.fetchone()
            admin_nav_badge_count = row["total"] if row else 0
        else:
            last_seen_user = seen_row["last_seen_user_bookings"] if seen_row else None
            if last_seen_user:
                cur.execute("""
                    SELECT COUNT(*) AS total FROM notifications
                    WHERE username = ? AND created_at > ?
                """, (session["user"], last_seen_user))
            else:
                cur.execute("""
                    SELECT COUNT(*) AS total FROM notifications
                    WHERE username = ?
                """, (session["user"],))
            row = cur.fetchone()
            user_nav_badge_count = row["total"] if row else 0

        conn.close()

    return {
        "unread_notification_count": int(unread_count),
        "user_nav_badge_count": int(user_nav_badge_count),
        "admin_nav_badge_count": int(admin_nav_badge_count)
    }


@app.route("/api/live-badges")
def api_live_badges():
    if not session.get("user"):
        return jsonify({"ok": False, "logged_in": False})
    payload = get_live_badge_payload()
    payload["ok"] = True
    payload["logged_in"] = True
    return jsonify(payload)

@app.before_request
def run_auto_cancel_before_each_request():
    auto_cancel_expired_bookings()
    remind_before_checkin()



def remind_before_checkin():
    conn = get_db()
    cur = conn.cursor()
    now_dt = get_now()
    upcoming_limit = now_dt + timedelta(days=1)

    cur.execute("""
        SELECT * FROM bookings
        WHERE status IN ('deposit_paid', 'remaining_counter_pending')
          AND checkin_reminder_sent_at IS NULL
    """)
    rows = cur.fetchall()

    changed_ids = []
    for booking in rows:
        try:
            checkin_dt = datetime.strptime(f"{booking['checkin']} {booking['checkin_time']}", "%Y-%m-%d %H:%M")
        except ValueError:
            continue

        if now_dt <= checkin_dt <= upcoming_limit:
            notify_user(
                booking["username"],
                "แจ้งเตือนก่อนวันเข้าพัก",
                f"รายการจองห้อง {booking['room_name']} ของคุณใกล้ถึงวันเข้าพักแล้ว และยังคงเหลือยอดชำระ {booking['remaining_due']} บาท กรุณาชำระส่วนที่เหลือก่อนเข้าพัก"
            )
            changed_ids.append(booking["id"])

    for booking_id in changed_ids:
        cur.execute("""
            UPDATE bookings
            SET checkin_reminder_sent_at = ?
            WHERE id = ?
        """, (now_dt.strftime("%Y-%m-%d %H:%M:%S"), booking_id))

    conn.commit()
    conn.close()


def auto_cancel_expired_bookings():
    conn = get_db()
    cur = conn.cursor()
    now_str = get_now().strftime("%Y-%m-%d %H:%M:%S")
    cur.execute("""
        SELECT * FROM bookings
        WHERE status = 'approved'
          AND payment_deadline IS NOT NULL
          AND payment_deadline < ?
    """, (now_str,))
    expired_bookings = cur.fetchall()

    if not expired_bookings:
        conn.close()
        return

    for booking in expired_bookings:
        cur.execute("UPDATE bookings SET status = 'cancelled' WHERE id = ?", (booking["id"],))

    conn.commit()
    conn.close()

    for booking in expired_bookings:
        notify_user(
            booking["username"],
            "การจองถูกยกเลิกอัตโนมัติ",
            f"รายการจองห้อง {booking['room_name']} ของคุณถูกยกเลิกอัตโนมัติ เพราะไม่ได้ชำระเงินภายในเวลาที่กำหนด",
            email_subject="การจองถูกยกเลิกอัตโนมัติ - โกเจมส์ ที่พักตากอากาศ",
            email_body=(
                f"รายการจองห้อง {booking['room_name']} ของคุณถูกยกเลิกอัตโนมัติ\n"
                f"เนื่องจากไม่ได้ชำระเงินภายในเวลาที่กำหนด ({booking['payment_deadline']})"
            )
        )


def init_db():
    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS users(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        first_name TEXT NOT NULL,
        last_name TEXT NOT NULL,
        nickname TEXT NOT NULL,
        gender TEXT NOT NULL,
        phone TEXT NOT NULL,
        email TEXT NOT NULL,
        username TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'user',
        last_seen_user_bookings TEXT,
        last_seen_admin_bookings TEXT
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS room_settings(
        room_id INTEGER PRIMARY KEY,
        is_open INTEGER NOT NULL DEFAULT 1
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS rooms(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        price REAL NOT NULL DEFAULT 0,
        detail TEXT,
        cover TEXT,
        images TEXT,
        amenities TEXT,
        is_open INTEGER NOT NULL DEFAULT 1,
        is_deleted INTEGER NOT NULL DEFAULT 0,
        status_text TEXT,
        maintenance_note TEXT,
        created_at TEXT,
        updated_at TEXT
    )
    """)

    room_columns = {row["name"] for row in cur.execute("PRAGMA table_info(rooms)").fetchall()}
    if "is_deleted" not in room_columns:
        cur.execute("ALTER TABLE rooms ADD COLUMN is_deleted INTEGER NOT NULL DEFAULT 0")
    if "status_text" not in room_columns:
        cur.execute("ALTER TABLE rooms ADD COLUMN status_text TEXT")
    if "maintenance_note" not in room_columns:
        cur.execute("ALTER TABLE rooms ADD COLUMN maintenance_note TEXT")
    if "created_at" not in room_columns:
        cur.execute("ALTER TABLE rooms ADD COLUMN created_at TEXT")
    if "updated_at" not in room_columns:
        cur.execute("ALTER TABLE rooms ADD COLUMN updated_at TEXT")

    cur.execute("SELECT COUNT(*) AS cnt FROM rooms")
    if (cur.fetchone()["cnt"] or 0) == 0:
        for room in ROOMS:
            cur.execute("""
                INSERT INTO rooms(id, name, price, detail, cover, images, amenities, is_open, is_deleted, status_text, maintenance_note, created_at, updated_at)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                room["id"],
                room.get("name", ""),
                float(room.get("price", 0)),
                room.get("detail", ""),
                room.get("cover", ""),
                json.dumps(room.get("images", [room.get("cover", "")]), ensure_ascii=False),
                json.dumps(room.get("amenities", []), ensure_ascii=False),
                1 if room.get("is_open", True) else 0,
                0,
                room.get("status_text", ""),
                room.get("maintenance_note", ""),
                db_now_str(),
                db_now_str()
            ))

    for room in ROOMS:
        cur.execute(
            "INSERT OR IGNORE INTO room_settings(room_id, is_open) VALUES (?, ?)",
            (room["id"], 1 if room.get("is_open", True) else 0)
        )

    cur.execute("""
    CREATE TABLE IF NOT EXISTS bookings(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT NOT NULL,
        room_id INTEGER NOT NULL,
        room_name TEXT NOT NULL,
        checkin TEXT NOT NULL,
        checkout TEXT NOT NULL,
        checkin_time TEXT NOT NULL,
        checkout_time TEXT NOT NULL,
        nights INTEGER NOT NULL,
        total REAL NOT NULL,
        paytype TEXT NOT NULL,
        paid REAL NOT NULL,
        status TEXT NOT NULL DEFAULT 'approved',
        created_at TEXT,
        approved_at TEXT,
        payment_deadline TEXT,
        slip_filename TEXT,
        batch_id TEXT,
        cancel_requested_from TEXT,
        cancel_requested_at TEXT,
        refund_completed_at TEXT,
        remaining_due REAL,
        checkin_reminder_sent_at TEXT,
        remaining_payment_method TEXT DEFAULT 'qr',
        checked_in_at TEXT,
        payment_method TEXT,
        receipt_no TEXT,
        receipt_issued_at TEXT,
        checked_out_at TEXT,
        counter_payment_method TEXT,
        note TEXT
    )
    """)

    booking_columns = {row["name"] for row in cur.execute("PRAGMA table_info(bookings)").fetchall()}
    if "batch_id" not in booking_columns:
        cur.execute("ALTER TABLE bookings ADD COLUMN batch_id TEXT")
    if "cancel_requested_from" not in booking_columns:
        cur.execute("ALTER TABLE bookings ADD COLUMN cancel_requested_from TEXT")
    if "cancel_requested_at" not in booking_columns:
        cur.execute("ALTER TABLE bookings ADD COLUMN cancel_requested_at TEXT")
    if "refund_completed_at" not in booking_columns:
        cur.execute("ALTER TABLE bookings ADD COLUMN refund_completed_at TEXT")
    if "refund_method" not in booking_columns:
        cur.execute("ALTER TABLE bookings ADD COLUMN refund_method TEXT")
    if "refund_bank" not in booking_columns:
        cur.execute("ALTER TABLE bookings ADD COLUMN refund_bank TEXT")
    if "refund_account" not in booking_columns:
        cur.execute("ALTER TABLE bookings ADD COLUMN refund_account TEXT")
    if "refund_amount" not in booking_columns:
        cur.execute("ALTER TABLE bookings ADD COLUMN refund_amount REAL DEFAULT 0")
    if "refund_note" not in booking_columns:
        cur.execute("ALTER TABLE bookings ADD COLUMN refund_note TEXT")
    if "cancelled_by" not in booking_columns:
        cur.execute("ALTER TABLE bookings ADD COLUMN cancelled_by TEXT")
    if "remaining_due" not in booking_columns:
        cur.execute("ALTER TABLE bookings ADD COLUMN remaining_due REAL DEFAULT 0")
    if "checkin_reminder_sent_at" not in booking_columns:
        cur.execute("ALTER TABLE bookings ADD COLUMN checkin_reminder_sent_at TEXT")
    if "remaining_payment_method" not in booking_columns:
        cur.execute("ALTER TABLE bookings ADD COLUMN remaining_payment_method TEXT DEFAULT 'qr'")
    if "checked_in_at" not in booking_columns:
        cur.execute("ALTER TABLE bookings ADD COLUMN checked_in_at TEXT")
    if "payment_method" not in booking_columns:
        cur.execute("ALTER TABLE bookings ADD COLUMN payment_method TEXT")
    if "receipt_no" not in booking_columns:
        cur.execute("ALTER TABLE bookings ADD COLUMN receipt_no TEXT")
    if "receipt_issued_at" not in booking_columns:
        cur.execute("ALTER TABLE bookings ADD COLUMN receipt_issued_at TEXT")
    if "checked_out_at" not in booking_columns:
        cur.execute("ALTER TABLE bookings ADD COLUMN checked_out_at TEXT")
    if "counter_payment_method" not in booking_columns:
        cur.execute("ALTER TABLE bookings ADD COLUMN counter_payment_method TEXT")
    if "note" not in booking_columns:
        cur.execute("ALTER TABLE bookings ADD COLUMN note TEXT")

    # ปรับฐานข้อมูลเก่าที่เคยใช้สถานะรอชำระเงิน ให้เข้ากับ Flow ใหม่: จองแล้วรอชำระเงินทันที
    cur.execute("UPDATE bookings SET status = 'approved' WHERE status = 'waiting_approval'")

    cur.execute("""
    CREATE TABLE IF NOT EXISTS room_blocks(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        room_id INTEGER NOT NULL,
        start_date TEXT NOT NULL,
        end_date TEXT NOT NULL,
        note TEXT,
        created_at TEXT NOT NULL,
        status TEXT NOT NULL DEFAULT 'blocked'
    )
    """)

    room_block_columns = {row["name"] for row in cur.execute("PRAGMA table_info(room_blocks)").fetchall()}
    if "status" not in room_block_columns:
        cur.execute("ALTER TABLE room_blocks ADD COLUMN status TEXT NOT NULL DEFAULT 'blocked'")

    cur.execute("""
    CREATE TABLE IF NOT EXISTS notifications(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT NOT NULL,
        title TEXT NOT NULL,
        message TEXT NOT NULL,
        is_read INTEGER NOT NULL DEFAULT 0,
        created_at TEXT NOT NULL
    )
    """)

    user_columns = {row["name"] for row in cur.execute("PRAGMA table_info(users)").fetchall()}
    if "last_seen_user_bookings" not in user_columns:
        cur.execute("ALTER TABLE users ADD COLUMN last_seen_user_bookings TEXT")
    if "last_seen_admin_bookings" not in user_columns:
        cur.execute("ALTER TABLE users ADD COLUMN last_seen_admin_bookings TEXT")
    if "last_seen_user_notification_id" not in user_columns:
        cur.execute("ALTER TABLE users ADD COLUMN last_seen_user_notification_id INTEGER DEFAULT 0")
    if "last_seen_admin_notification_id" not in user_columns:
        cur.execute("ALTER TABLE users ADD COLUMN last_seen_admin_notification_id INTEGER DEFAULT 0")

    cur.execute("SELECT id FROM users WHERE username = ?", ("admin",))
    if cur.fetchone() is None:
        cur.execute("""
            INSERT INTO users(
                first_name, last_name, nickname, gender, phone, email,
                username, password, role
            )
            VALUES(?,?,?,?,?,?,?,?,?)
        """, (
            "Admin", "System", "admin", "อื่นๆ", "0000000000",
            "admin@email.com", "admin", "admin123", "admin"
        ))

    conn.commit()
    conn.close()


def get_room(room_id):
    for room in get_rooms_with_status():
        if room["id"] == room_id:
            return room
    return None


def create_qr(booking_id, amount):
    payload = generate_promptpay_payload(PROMPTPAY_ID, amount)
    img = qrcode.make(payload)
    filename = f"qr_{booking_id}.png"
    path = os.path.join(STATIC_DIR, filename)
    img.save(path)
    return filename


def is_overlap(start1, end1, start2, end2):
    return start1 < end2 and start2 < end1


def block_overlaps_booking(booking_start, booking_end, block_start, block_end):
    block_start_dt = datetime.combine(block_start, datetime.min.time())
    block_end_dt = datetime.combine(block_end + timedelta(days=1), datetime.min.time())
    return is_overlap(booking_start, booking_end, block_start_dt, block_end_dt)


def calculate_charge_nights(checkin_dt, checkout_dt):
    day_diff = (checkout_dt.date() - checkin_dt.date()).days
    return max(1, day_diff)


def calculate_due_now(total_amount, paytype):
    return float(total_amount) if paytype == "full" else round(float(total_amount) * 0.25, 2)


def calculate_remaining_due(total_amount, paytype):
    due_now = calculate_due_now(total_amount, paytype)
    return round(float(total_amount) - float(due_now), 2)


def calculate_refund_amount(booking, cancelled_by="user"):
    paid = float(booking["paid"] or 0)
    paytype = booking["paytype"]
    if cancelled_by == "admin":
        return round(paid, 2)
    if paytype == "full":
        return round(paid * 0.5, 2)
    return 0.0




def choose_group_status(rows):
    statuses = [row["status"] for row in rows if row["status"]]
    if not statuses:
        return ""
    if len(set(statuses)) == 1:
        return statuses[0]

    priority = [
        "checked_in",
        "cancel_requested",
        "refund_pending",
        "refunded",
        "waiting_remaining_slip_approval",
        "remaining_counter_pending",
        "deposit_paid",
        "waiting_slip_approval",
        "approved",
        "waiting_approval",
        "fully_paid",
        "checked_out",
        "no_show",
        "rejected",
        "cancelled",
    ]
    for status in priority:
        if status in statuses:
            return status
    return statuses[0]


def build_booking_groups(rows):
    grouped = {}
    for row in rows:
        key = row["batch_id"] if row["batch_id"] else f"single-{row['id']}"
        grouped.setdefault(key, []).append(row)

    countable_statuses = {
        "waiting_approval",
        "approved",
        "waiting_slip_approval",
        "deposit_paid",
        "waiting_remaining_slip_approval",
        "remaining_counter_pending",
        "fully_paid",
        "checked_in",
    }

    groups = []
    for key, group_rows in grouped.items():
        ordered = sorted(
            group_rows,
            key=lambda r: datetime.strptime(f"{r['checkin']} {r['checkin_time']}", "%Y-%m-%d %H:%M")
        )
        representative = max(group_rows, key=lambda r: r["id"])
        active_rows = [r for r in ordered if r["status"] in countable_statuses]
        rows_for_totals = active_rows if active_rows else ordered
        first_row = rows_for_totals[0]
        last_row = max(
            rows_for_totals,
            key=lambda r: datetime.strptime(f"{r['checkout']} {r['checkout_time']}", "%Y-%m-%d %H:%M")
        )
        deadlines = [r["payment_deadline"] for r in rows_for_totals if r["payment_deadline"]]
        slip_row = next((r for r in ordered if r["slip_filename"]), representative)

        segments = []
        for r in ordered:
            segments.append({
                "id": r["id"],
                "range_text": f"{r['checkin']} {r['checkin_time']} → {r['checkout']} {r['checkout_time']}",
                "status": r["status"],
                "nights": int(float(r["nights"] or 0)),
                "total": round(float(r["total"] or 0), 2),
                "remaining_due": round(float(r["remaining_due"] or 0), 2),
                "cancel_allowed": r["status"] in {
                    "waiting_approval", "approved", "waiting_slip_approval",
                    "deposit_paid", "remaining_counter_pending", "fully_paid"
                }
            })

        group = dict(representative)
        group.update({
            "id": representative["id"],
            "display_id": representative["id"],
            "action_id": representative["id"],
            "room_name": representative["room_name"],
            "checkin": first_row["checkin"],
            "checkin_time": first_row["checkin_time"],
            "checkout": last_row["checkout"],
            "checkout_time": last_row["checkout_time"],
            "nights": sum(float(r["nights"] or 0) for r in rows_for_totals),
            "total": round(sum(float(r["total"] or 0) for r in rows_for_totals), 2),
            "paid": round(sum(float(r["paid"] or 0) for r in rows_for_totals), 2),
            "remaining_due": round(sum(float(r["remaining_due"] or 0) for r in rows_for_totals), 2),
            "status": choose_group_status(rows_for_totals if rows_for_totals else group_rows),
            "payment_deadline": min(deadlines) if deadlines else None,
            "slip_filename": slip_row["slip_filename"],
            "ranges_text": [segment["range_text"] for segment in segments],
            "segments": segments,
            "segment_count": len(group_rows),
            "active_segment_count": len(active_rows),
            "is_grouped": len(group_rows) > 1,
            "sort_id": representative["id"],
        })
        groups.append(group)

    groups.sort(key=lambda r: r["sort_id"], reverse=True)
    return groups


def get_today_available_time_options():
    """เวลาเข้าพักสำหรับลูกค้าในกรณีเลือกเข้าพักวันนี้
    - ปกติเริ่ม 14:00 น.
    - ถ้าเวลาปัจจุบันเลย 14:00 แล้ว ให้เริ่มจากชั่วโมงถัดไป
    """
    now = datetime.now()
    start_hour = max(14, now.hour + (1 if now.minute > 0 or now.second > 0 or now.microsecond > 0 else 0))
    if start_hour > 23:
        return []
    return [f"{hour:02d}:00" for hour in range(start_hour, 24)]


def get_all_time_options():
    return [f"{hour:02d}:00" for hour in range(24)]


def get_checkin_time_options():
    # ลูกค้าจองล่วงหน้าเลือกเวลาเข้าพักได้ตั้งแต่ 14:00 น. เป็นต้นไป
    return [f"{hour:02d}:00" for hour in range(14, 24)]


def get_checkout_time_options():
    return get_all_time_options()


def get_calendar_data(days=14, start_day=None):
    conn = get_db()
    cur = conn.cursor()

    if start_day is None:
        start_day = date.today()
    day_list = [start_day + timedelta(days=i) for i in range(days)]

    placeholders = ",".join("?" for _ in BOOKED_STATUSES)
    cur.execute(f"""
        SELECT bookings.*, users.first_name, users.last_name
        FROM bookings
        LEFT JOIN users ON bookings.username = users.username
        WHERE bookings.status IN ({placeholders})
    """, BOOKED_STATUSES)
    active_bookings = cur.fetchall()

    cur.execute("SELECT * FROM room_blocks ORDER BY start_date ASC")
    blocks = cur.fetchall()
    conn.close()

    today = date.today()

    room_rows = []
    for room in get_rooms_with_status():
        cells = []
        room_closed = not room.get("is_open", True)

        for day in day_list:
            state = "free"
            detail = None

            if room_closed:
                state = "blocked"
                detail = {
                    "guest_name": "-",
                    "room_name": room["name"],
                    "checkin": day.strftime("%Y-%m-%d"),
                    "checkin_time": "-",
                    "checkout": day.strftime("%Y-%m-%d"),
                    "checkout_time": "-",
                    "status_text": "ปิดปรับปรุง",
                    "note": room.get("maintenance_note") or "-",
                    "source": "room_closed"
                }
            else:
                for block in blocks:
                    if block["room_id"] == room["id"]:
                        block_start = datetime.strptime(block["start_date"], "%Y-%m-%d").date()
                        block_end = datetime.strptime(block["end_date"], "%Y-%m-%d").date()
                        if block_start <= day <= block_end:
                            block_status = block["status"] if "status" in block.keys() and block["status"] else "blocked"
                            if block_status not in MANUAL_DAY_STATUSES:
                                block_status = "blocked"
                            state = block_status
                            detail = {
                                "guest_name": "-",
                                "room_name": room["name"],
                                "checkin": block["start_date"],
                                "checkin_time": "-",
                                "checkout": block["end_date"],
                                "checkout_time": "-",
                                "status_text": MANUAL_DAY_STATUSES.get(block_status, "ไม่ว่าง"),
                                "note": block["note"] if block["note"] else "-",
                                "source": "manual"
                            }
                            break

                if state == "free":
                    for booking in active_bookings:
                        if booking["room_id"] == room["id"]:
                            booking_start = datetime.strptime(f"{booking['checkin']} {booking['checkin_time']}", "%Y-%m-%d %H:%M")
                            booking_end = datetime.strptime(f"{booking['checkout']} {booking['checkout_time']}", "%Y-%m-%d %H:%M")
                            day_start = datetime.combine(day, datetime.min.time())
                            day_end = day_start + timedelta(days=1)
                            if is_overlap(booking_start, booking_end, day_start, day_end):
                                if day == today and booking_start <= datetime.now() < booking_end:
                                    state = "occupied"
                                    status_text = "มีผู้พักอยู่"
                                else:
                                    state = "booked"
                                    status_text = "มีผู้จอง"

                                full_name = " ".join(
                                    x for x in [booking["first_name"], booking["last_name"]] if x
                                ).strip()
                                detail = {
                                    "guest_name": full_name if full_name else booking["username"],
                                    "room_name": booking["room_name"],
                                    "checkin": booking["checkin"],
                                    "checkin_time": booking["checkin_time"],
                                    "checkout": booking["checkout"],
                                    "checkout_time": booking["checkout_time"],
                                    "status_text": status_text,
                                    "note": f"รหัสการจอง #{booking['id']}",
                                    "source": "booking",
                                    "booking_id": booking["id"]
                                }
                                break

            cells.append({
                "date": day,
                "date_text": day.strftime("%Y-%m-%d"),
                "state": state,
                "detail": detail,
                "room_id": room["id"],
                "room_name": room["name"]
            })
        room_rows.append({"room": room, "cells": cells})

    return day_list, room_rows


@app.route("/")
def index():
    auto_cancel_expired_bookings()
    return render_template("index.html", rooms=get_rooms_with_status())


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        first_name = request.form["first_name"].strip()
        last_name = request.form["last_name"].strip()
        nickname = request.form["nickname"].strip()
        gender = request.form["gender"].strip()
        phone = request.form["phone"].strip()
        email = request.form["email"].strip()
        username = request.form["username"].strip()
        password = request.form["password"].strip()
        confirm_password = request.form["confirm_password"].strip()

        if not all([first_name, last_name, nickname, gender, phone, email, username, password, confirm_password]):
            flash("กรอกข้อมูลให้ครบ")
            return redirect("/register")

        if not phone.isdigit() or len(phone) != 10 or not phone.startswith("0"):
            flash("กรุณากรอกเบอร์โทรศัพท์เป็นตัวเลข 10 หลัก และขึ้นต้นด้วย 0")
            return redirect("/register")

        if not re.fullmatch(r"[A-Za-z0-9_.]{3,30}", username):
            flash("ชื่อผู้ใช้ต้องเป็นภาษาอังกฤษ ตัวเลข _ หรือ . เท่านั้น และมีความยาว 3-30 ตัวอักษร")
            return redirect("/register")

        if len(password) < 8:
            flash("รหัสผ่านต้องมีอย่างน้อย 8 ตัวอักษร")
            return redirect("/register")

        if password != confirm_password:
            flash("รหัสผ่านและยืนยันรหัสผ่านไม่ตรงกัน")
            return redirect("/register")

        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT id FROM users WHERE username = ?", (username,))
        if cur.fetchone():
            conn.close()
            flash("ชื่อผู้ใช้นี้มีอยู่แล้ว")
            return redirect("/register")

        cur.execute("""
            INSERT INTO users(first_name, last_name, nickname, gender, phone, email, username, password, role)
            VALUES(?,?,?,?,?,?,?,?,?)
        """, (first_name, last_name, nickname, gender, phone, email, username, password, "user"))
        conn.commit()
        conn.close()

        flash("สมัครสมาชิกสำเร็จ กรุณาเข้าสู่ระบบ")
        return redirect("/login")

    return render_template("register.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["password"].strip()

        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT * FROM users WHERE username = ? AND password = ?", (username, password))
        user = cur.fetchone()
        conn.close()

        if user:
            session["user"] = user["username"]
            session["role"] = user["role"]
            flash("เข้าสู่ระบบสำเร็จ")
            return redirect("/")
        else:
            flash("ชื่อผู้ใช้หรือรหัสผ่านไม่ถูกต้อง")

    return render_template("login.html")


@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    if request.method == "POST":
        email = request.form["email"].strip()

        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT * FROM users WHERE email = ? ORDER BY id DESC LIMIT 1", (email,))
        user = cur.fetchone()
        conn.close()

        if not user:
            flash("ไม่พบบัญชีผู้ใช้งานที่ใช้อีเมลนี้ กรุณาตรวจสอบอีเมลอีกครั้ง หรือสมัครสมาชิกหากยังไม่มีบัญชี")
            return redirect("/forgot-password")

        sent = send_reset_password_email(user)
        if sent:
            flash("ส่งลิงก์รีเซ็ตรหัสผ่านไปยังอีเมลของคุณแล้ว กรุณาตรวจสอบกล่องจดหมายหรือโฟลเดอร์สแปม ลิงก์จะหมดอายุภายใน 30 นาที")
            return redirect("/login")

        flash("ไม่สามารถส่งอีเมลรีเซ็ตรหัสผ่านได้ในขณะนี้ กรุณาลองใหม่อีกครั้ง หรือติดต่อผู้ดูแลระบบ")
        return redirect("/forgot-password")

    return render_template("forgot_password.html")


@app.route("/reset-password/<token>", methods=["GET", "POST"])
def reset_password(token):
    try:
        payload = serializer.loads(token, salt="reset-password", max_age=1800)
    except SignatureExpired:
        flash("ลิงก์รีเซ็ตรหัสผ่านหมดอายุแล้ว กรุณาขอใหม่อีกครั้ง")
        return redirect("/forgot-password")
    except BadSignature:
        flash("ลิงก์รีเซ็ตรหัสผ่านไม่ถูกต้อง")
        return redirect("/forgot-password")

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM users WHERE id = ? AND email = ?", (payload["user_id"], payload["email"]))
    user = cur.fetchone()

    if not user:
        conn.close()
        flash("ไม่พบบัญชีผู้ใช้งาน")
        return redirect("/forgot-password")

    if request.method == "POST":
        password = request.form["password"].strip()
        confirm_password = request.form["confirm_password"].strip()

        if not password or not confirm_password:
            conn.close()
            flash("กรุณากรอกรหัสผ่านให้ครบ")
            return redirect(request.url)

        if password != confirm_password:
            conn.close()
            flash("รหัสผ่านและยืนยันรหัสผ่านไม่ตรงกัน")
            return redirect(request.url)

        cur.execute("UPDATE users SET password = ? WHERE id = ?", (password, user["id"]))
        conn.commit()
        conn.close()

        flash("เปลี่ยนรหัสผ่านสำเร็จ กรุณาเข้าสู่ระบบ")
        return redirect("/login")

    conn.close()
    return render_template("reset_password.html", token=token)



@app.route("/profile", methods=["GET", "POST"])
def profile():
    if "user" not in session:
        flash("กรุณาเข้าสู่ระบบก่อน")
        return redirect("/login")

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM users WHERE username = ?", (session["user"],))
    user = cur.fetchone()

    if not user:
        conn.close()
        flash("ไม่พบบัญชีผู้ใช้")
        return redirect("/logout")

    if request.method == "POST":
        first_name = request.form.get("first_name", "").strip()
        last_name = request.form.get("last_name", "").strip()
        nickname = request.form.get("nickname", "").strip()
        gender = request.form.get("gender", "").strip()
        phone = request.form.get("phone", "").strip()
        email = request.form.get("email", "").strip()

        if not all([first_name, last_name, nickname, gender, phone, email]):
            conn.close()
            flash("กรุณากรอกข้อมูลโปรไฟล์ให้ครบ")
            return redirect("/profile")

        if not phone.isdigit() or len(phone) != 10 or not phone.startswith("0"):
            conn.close()
            flash("กรุณากรอกเบอร์โทรศัพท์เป็นตัวเลข 10 หลัก และขึ้นต้นด้วย 0")
            return redirect("/profile")

        cur.execute("""
            UPDATE users
            SET first_name = ?, last_name = ?, nickname = ?, gender = ?, phone = ?, email = ?
            WHERE username = ?
        """, (first_name, last_name, nickname, gender, phone, email, session["user"]))
        conn.commit()
        conn.close()
        flash("บันทึกโปรไฟล์เรียบร้อยแล้ว")
        return redirect("/profile")

    conn.close()
    return render_template("profile.html", user=user)


@app.route("/profile/change-password", methods=["POST"])
def profile_change_password():
    if "user" not in session:
        flash("กรุณาเข้าสู่ระบบก่อน")
        return redirect("/login")

    current_password = request.form.get("current_password", "").strip()
    new_password = request.form.get("new_password", "").strip()
    confirm_password = request.form.get("confirm_password", "").strip()

    if not current_password or not new_password or not confirm_password:
        flash("กรุณากรอกรหัสผ่านให้ครบ")
        return redirect("/profile")

    if new_password != confirm_password:
        flash("รหัสผ่านใหม่และยืนยันรหัสผ่านไม่ตรงกัน")
        return redirect("/profile")

    if len(new_password) < 8:
        flash("รหัสผ่านใหม่ต้องมีอย่างน้อย 8 ตัวอักษร")
        return redirect("/profile")

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM users WHERE username = ?", (session["user"],))
    user = cur.fetchone()

    if not user or user["password"] != current_password:
        conn.close()
        flash("รหัสผ่านเดิมไม่ถูกต้อง")
        return redirect("/profile")

    cur.execute("UPDATE users SET password = ? WHERE username = ?", (new_password, session["user"]))
    conn.commit()
    conn.close()

    flash("เปลี่ยนรหัสผ่านเรียบร้อยแล้ว")
    return redirect("/profile")

@app.route("/logout")
def logout():
    session.clear()
    flash("ออกจากระบบแล้ว")
    return redirect("/")


@app.route("/booking/<int:room_id>", methods=["GET", "POST"])
def booking(room_id):
    auto_cancel_expired_bookings()

    if "user" not in session:
        flash("กรุณาเข้าสู่ระบบก่อนจอง")
        return redirect("/login")

    if session.get("role") == "admin" and request.method == "POST":
        flash("ผู้ดูแลระบบไม่สามารถทำรายการจองผ่านหน้านี้ได้")
        return redirect(url_for("booking", room_id=room_id))

    room = get_room(room_id)
    if room is None:
        return "ไม่พบห้องพัก", 404
    if room.get("is_open", True) is False:
        flash(room.get("maintenance_note", "ห้องนี้ยังไม่พร้อมให้บริการ"))
        return redirect("/")

    all_time_options = get_all_time_options()
    today_time_options = get_today_available_time_options()
    checkin_time_options = get_checkin_time_options()
    checkout_time_options = get_checkout_time_options()

    default_checkin_date = date.today()
    default_checkout_date = default_checkin_date + timedelta(days=1)

    if request.method == "POST":
        paytype = request.form["paytype"]
        raw_extra_segments = request.form.get("extra_segments", "[]")

        try:
            extra_segments = json.loads(raw_extra_segments) if raw_extra_segments else []
        except json.JSONDecodeError:
            extra_segments = []

        segment_payloads = [{
            "checkin": request.form["checkin"],
            "checkout": request.form["checkout"],
            "checkin_time": request.form["checkin_time"],
            "checkout_time": RESORT_CHECKOUT_GUIDE
        }]

        for seg in extra_segments:
            if isinstance(seg, dict):
                segment_payloads.append({
                    "checkin": str(seg.get("checkin", "")).strip(),
                    "checkout": str(seg.get("checkout", "")).strip(),
                    "checkin_time": str(seg.get("checkin_time", "")).strip(),
                    "checkout_time": RESORT_CHECKOUT_GUIDE
                })

        validated_segments = []
        now = get_now()

        for idx, seg in enumerate(segment_payloads, start=1):
            checkin = seg["checkin"]
            checkout = seg["checkout"]
            checkin_time = seg["checkin_time"]
            checkout_time = RESORT_CHECKOUT_GUIDE

            try:
                d1 = datetime.strptime(checkin, "%Y-%m-%d").date()
                d2 = datetime.strptime(checkout, "%Y-%m-%d").date()
                dt1 = datetime.strptime(f"{checkin} {checkin_time}", "%Y-%m-%d %H:%M")
                dt2 = datetime.strptime(f"{checkout} {checkout_time}", "%Y-%m-%d %H:%M")
            except ValueError:
                flash(f"ช่วงที่ {idx}: รูปแบบวันที่หรือเวลาไม่ถูกต้อง")
                return redirect(url_for("booking", room_id=room_id))

            if d1 < date.today():
                flash(f"ช่วงที่ {idx}: ไม่สามารถจองย้อนหลังได้")
                return redirect(url_for("booking", room_id=room_id))

            # ลูกค้าจองผ่านหน้าเว็บ: เวลาเข้าพักมาตรฐานเริ่ม 14:00 น.
            # ถ้าเลือกเข้าพักวันนี้ ระบบจะตัดเวลาที่ผ่านมาแล้วออก และบังคับเป็นชั่วโมงถัดไป
            if d1 == date.today():
                min_hour = max(14, now.hour + (1 if now.minute > 0 or now.second > 0 or now.microsecond > 0 else 0))
            else:
                min_hour = 14

            try:
                selected_hour = int(checkin_time.split(":", 1)[0])
            except Exception:
                flash(f"ช่วงที่ {idx}: เวลาเข้าพักไม่ถูกต้อง")
                return redirect(url_for("booking", room_id=room_id))

            if min_hour > 23:
                flash(f"ช่วงที่ {idx}: วันนี้ไม่สามารถเลือกเวลาเข้าพักได้แล้ว กรุณาเลือกวันถัดไป")
                return redirect(url_for("booking", room_id=room_id))

            if selected_hour < min_hour:
                flash(f"ช่วงที่ {idx}: เวลาเข้าพักต้องเริ่มตั้งแต่ {min_hour:02d}:00 น. เป็นต้นไป")
                return redirect(url_for("booking", room_id=room_id))

            if dt2 <= dt1:
                flash(f"ช่วงที่ {idx}: วันและเวลาออกต้องมากกว่าวันและเวลาเข้า")
                return redirect(url_for("booking", room_id=room_id))

            validated_segments.append({
                "checkin": checkin,
                "checkout": checkout,
                "checkin_time": checkin_time,
                "checkout_time": checkout_time,
                "start_dt": dt1,
                "end_dt": dt2,
                "nights": calculate_charge_nights(dt1, dt2)
            })

        for i in range(len(validated_segments)):
            for j in range(i + 1, len(validated_segments)):
                if is_overlap(
                    validated_segments[i]["start_dt"],
                    validated_segments[i]["end_dt"],
                    validated_segments[j]["start_dt"],
                    validated_segments[j]["end_dt"]
                ):
                    flash("ช่วงวันเข้าพักที่เพิ่มมาซ้อนกันเอง กรุณาตรวจสอบอีกครั้ง")
                    return redirect(url_for("booking", room_id=room_id))

        conn = get_db()
        cur = conn.cursor()

        active_statuses = BOOKED_STATUSES
        placeholders = ",".join("?" for _ in active_statuses)

        # กันลูกค้าคนเดิมจองช่วงเวลาซ้ำกับรายการเดิมของตัวเอง
        cur.execute(f"""
            SELECT * FROM bookings
            WHERE username = ?
              AND status IN ({placeholders})
        """, (session["user"], *active_statuses))
        existing_user_bookings = cur.fetchall()

        for seg in validated_segments:
            for existing in existing_user_bookings:
                exist_in = datetime.strptime(f"{existing['checkin']} {existing['checkin_time']}", "%Y-%m-%d %H:%M")
                exist_out = datetime.strptime(f"{existing['checkout']} {existing['checkout_time']}", "%Y-%m-%d %H:%M")
                if is_overlap(seg["start_dt"], seg["end_dt"], exist_in, exist_out):
                    conn.close()
                    flash(
                        f"ช่วงวันที่ {seg['checkin']} {seg['checkin_time']} ถึง {seg['checkout']} {seg['checkout_time']} "
                        f"ซ้ำกับรายการเดิมของคุณ (ห้อง {existing['room_name']})"
                    )
                    return redirect(url_for("booking", room_id=room_id))

        # กันจองห้องซ้ำกับลูกค้าคนอื่นหรือรายการเดิมของห้องเดียวกัน
        cur.execute(f"""
            SELECT * FROM bookings
            WHERE room_id = ?
              AND status IN ({placeholders})
        """, (room_id, *active_statuses))
        existing_room_bookings = cur.fetchall()

        for seg in validated_segments:
            for existing in existing_room_bookings:
                exist_in = datetime.strptime(f"{existing['checkin']} {existing['checkin_time']}", "%Y-%m-%d %H:%M")
                exist_out = datetime.strptime(f"{existing['checkout']} {existing['checkout_time']}", "%Y-%m-%d %H:%M")
                if is_overlap(seg["start_dt"], seg["end_dt"], exist_in, exist_out):
                    conn.close()
                    flash(
                        f"ห้องนี้ถูกจองแล้วในช่วง {existing['checkin']} {existing['checkin_time']} "
                        f"ถึง {existing['checkout']} {existing['checkout_time']} กรุณาเลือกช่วงเวลาอื่น"
                    )
                    return redirect(url_for("booking", room_id=room_id))

        # กันจองทับช่วงที่ผู้ดูแลระบบบล็อก/ปิดห้องไว้
        cur.execute("""
            SELECT * FROM room_blocks
            WHERE room_id = ?
              AND COALESCE(status, 'blocked') != 'free'
        """, (room_id,))
        room_blocks = cur.fetchall()

        for seg in validated_segments:
            for block in room_blocks:
                try:
                    block_start = datetime.strptime(block["start_date"], "%Y-%m-%d").date()
                    block_end = datetime.strptime(block["end_date"], "%Y-%m-%d").date()
                except Exception:
                    continue
                if block_overlaps_booking(seg["start_dt"], seg["end_dt"], block_start, block_end):
                    conn.close()
                    flash(
                        f"ห้องนี้ไม่พร้อมให้จองในช่วง {block['start_date']} ถึง {block['end_date']} "
                        "กรุณาเลือกช่วงเวลาอื่น"
                    )
                    return redirect(url_for("booking", room_id=room_id))

        batch_id = uuid.uuid4().hex
        created_count = 0

        for seg in validated_segments:
            total = seg["nights"] * room["price"]
            paid = calculate_due_now(total, paytype)
            remaining_due = calculate_remaining_due(total, paytype)

            cur.execute("""
                INSERT INTO bookings(
                    username, room_id, room_name, checkin, checkout,
                    checkin_time, checkout_time, nights, total, paytype,
                    paid, status, created_at, batch_id, remaining_due
                )
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                session["user"], room_id, room["name"], seg["checkin"], seg["checkout"],
                seg["checkin_time"], seg["checkout_time"], seg["nights"], total, paytype,
                paid, "approved", get_now().strftime("%Y-%m-%d %H:%M:%S"), batch_id,
                remaining_due
            ))
            created_count += 1

        conn.commit()
        conn.close()

        notify_admin(
            "มีรายการจองใหม่",
            f"ลูกค้า {session['user']} จองห้อง {room['name']} จำนวน {created_count} รายการ และอยู่ในสถานะรอชำระเงิน",
            email_subject="มีรายการจองใหม่ - GoJames Vacation Home",
            email_body=(
                f"มีรายการจองใหม่จากลูกค้า {session['user']}\n"
                f"ห้องพัก: {room['name']}\n"
                f"จำนวนรายการ: {created_count}\n"
                "สถานะ: รอชำระเงิน"
            )
        )
        notify_user(
            session["user"],
            "จองห้องพักเรียบร้อย - รอชำระเงิน",
            f"ระบบได้รับรายการจองห้อง {room['name']} แล้ว กรุณาดำเนินการชำระเงิน",
            email_subject="จองห้องพักเรียบร้อย - กรุณาชำระเงิน",
            email_body=f"ระบบได้รับรายการจองห้อง {room['name']} แล้ว กรุณาดำเนินการชำระเงิน"
        )

        if created_count > 1:
            flash(f"จองห้องเรียบร้อยแล้ว จำนวน {created_count} ช่วงวัน กรุณาดำเนินการชำระเงิน")
        else:
            flash("จองห้องเรียบร้อยแล้ว กรุณาดำเนินการชำระเงิน")
        return redirect("/my-bookings")

    return render_template(
        "booking.html",
        room=room,
        all_time_options=all_time_options,
        today_time_options=today_time_options,
        checkin_time_options=checkin_time_options,
        checkout_time_options=checkout_time_options,
        default_checkin_str=default_checkin_date.strftime("%Y-%m-%d"),
        default_checkout_str=default_checkout_date.strftime("%Y-%m-%d"),
        standard_checkin_time=RESORT_CHECKIN_GUIDE,
        standard_checkout_time=RESORT_CHECKOUT_GUIDE,
        today_str=date.today().strftime("%Y-%m-%d")
    )


@app.route("/my-bookings")
def my_bookings():
    auto_cancel_expired_bookings()

    if "user" not in session:
        flash("กรุณาเข้าสู่ระบบก่อน")
        return redirect("/login")

    conn = get_db()
    cur = conn.cursor()

    # เปิดหน้า "การจองของฉัน" แล้วให้ badge ของเมนูนี้หาย
    cur.execute("SELECT COALESCE(MAX(id), 0) AS max_id FROM notifications WHERE username = ?", (session["user"],))
    latest_notification = cur.fetchone()
    latest_notification_id = latest_notification["max_id"] if latest_notification else 0

    cur.execute("""
        UPDATE users
        SET last_seen_user_bookings = ?, last_seen_user_notification_id = ?
        WHERE username = ?
    """, (db_now_str(), latest_notification_id, session["user"]))

    cur.execute("SELECT * FROM bookings WHERE username = ? ORDER BY id DESC", (session["user"],))
    bookings = cur.fetchall()
    grouped_bookings = build_booking_groups(bookings)
    conn.commit()
    conn.close()

    return render_template("my_bookings.html", bookings=grouped_bookings)




@app.route("/my-bookings/check-out/<int:booking_id>", methods=["POST"])
def user_check_out(booking_id):
    """ปิดการเช็กเอาท์ด้วยตนเองของลูกค้า ให้ผู้ดูแลระบบ/หน้าเคาน์เตอร์เป็นผู้จัดการเท่านั้น"""
    if "user" not in session:
        flash("กรุณาเข้าสู่ระบบก่อน")
        return redirect("/login")
    flash("หากต้องการเช็กเอาท์ กรุณาติดต่อหน้าเคาน์เตอร์ เพื่อให้ผู้ดูแลดำเนินการในระบบ")
    return redirect("/my-bookings")


@app.route("/cancel/<int:booking_id>", methods=["GET", "POST"])
def cancel_booking(booking_id):
    if "user" not in session:
        flash("กรุณาเข้าสู่ระบบก่อน")
        return redirect("/login")

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM bookings WHERE id = ?", (booking_id,))
    booking = cur.fetchone()

    if booking is None:
        conn.close()
        flash("ไม่พบรายการจอง")
        return redirect("/my-bookings")

    is_admin = session.get("role") == "admin"
    is_owner = booking["username"] == session.get("user")
    if not is_admin and not is_owner:
        conn.close()
        flash("คุณไม่มีสิทธิ์ยกเลิกรายการนี้")
        return redirect("/my-bookings")

    current_status = booking["status"]
    now_str = get_now().strftime("%Y-%m-%d %H:%M:%S")
    immediate_cancel_statuses = {"waiting_approval", "approved", "waiting_slip_approval"}
    paid_statuses = {"deposit_paid", "remaining_counter_pending", "fully_paid"}

    if current_status in immediate_cancel_statuses:
        cur.execute(
            """
            UPDATE bookings
            SET status = 'cancelled',
                cancel_requested_from = NULL,
                cancel_requested_at = NULL,
                refund_completed_at = NULL,
                refund_amount = 0,
                refund_method = NULL,
                refund_bank = NULL,
                refund_account = NULL,
                refund_note = NULL,
                cancelled_by = ?
            WHERE id = ?
            """,
            ("admin" if is_admin else "user", booking_id)
        )
        conn.commit()
        conn.close()
        if is_admin:
            notify_user(
                booking["username"],
                "รายการจองถูกยกเลิก",
                f"รายการจองห้อง {booking['room_name']} ถูกยกเลิกโดยผู้ดูแลระบบ"
            )
        else:
            notify_admin(
                "ลูกค้ายกเลิกการจอง",
                f"ลูกค้า {booking['username']} ยกเลิกรายการจองห้อง {booking['room_name']}"
            )
        flash("ยกเลิกการจองเรียบร้อย")
        return redirect("/admin" if is_admin else "/my-bookings")

    if current_status in paid_statuses:
        if is_admin:
            refund_amount = calculate_refund_amount(booking, "admin")
            cur.execute(
                """
                UPDATE bookings
                SET status = 'refund_pending',
                    cancel_requested_from = ?,
                    cancel_requested_at = ?,
                    refund_amount = ?,
                    refund_note = ?,
                    cancelled_by = 'admin',
                    refund_method = CASE WHEN paytype = 'deposit' THEN 'refund' ELSE refund_method END
                WHERE id = ?
                """,
                (current_status, now_str, refund_amount, "ผู้ดูแลระบบเป็นผู้ยกเลิกรายการ คืนเงิน 100%", booking_id)
            )
            conn.commit()
            conn.close()
            notify_user(
                booking["username"],
                "รายการจองถูกยกเลิกโดยผู้ดูแลระบบ",
                f"รายการจองห้อง {booking['room_name']} ถูกยกเลิกโดยผู้ดูแลระบบ และอยู่ระหว่างดำเนินการคืนเงิน {refund_amount:.2f} บาท"
            )
            flash(f"ยกเลิกรายการแล้ว และต้องคืนเงิน {refund_amount:.2f} บาท")
            return redirect("/admin")

        if request.method == "POST":
            refund_method = request.form.get("refund_method", "").strip()
            refund_account = request.form.get("refund_account", "").strip()
            refund_bank = request.form.get("refund_bank", "").strip()

            valid, final_bank, normalized_account, error_message = validate_refund_destination(
                refund_method, refund_account, refund_bank
            )
            if not valid:
                conn.close()
                flash(error_message)
                return redirect(request.path)

            refund_amount = calculate_refund_amount(booking, "user")
            note = "ลูกค้ายกเลิกหลังชำระเงิน: จ่ายเต็มคืน 50% / มัดจำไม่คืน"
            cur.execute(
                """
                UPDATE bookings
                SET status = 'cancel_requested',
                    cancel_requested_from = ?,
                    cancel_requested_at = ?,
                    refund_method = ?,
                    refund_bank = ?,
                    refund_account = ?,
                    refund_amount = ?,
                    refund_note = ?,
                    cancelled_by = 'user'
                WHERE id = ?
                """,
                (
                    current_status,
                    now_str,
                    refund_method.lower(),
                    final_bank,
                    normalized_account,
                    refund_amount,
                    note,
                    booking_id
                )
            )
            conn.commit()
            conn.close()
            notify_user(
                booking["username"],
                "ส่งคำขอยกเลิกแล้ว",
                f"ระบบได้รับคำขอยกเลิกห้อง {booking['room_name']} แล้ว กรุณารอผู้ดูแลระบบตรวจสอบ"
            )
            notify_admin(
                "มีคำขอยกเลิกการจอง",
                f"ลูกค้า {booking['username']} ส่งคำขอยกเลิกรายการจองห้อง {booking['room_name']} ยอดคืนเบื้องต้น {refund_amount:.2f} บาท",
                email_subject="มีคำขอยกเลิกการจอง - GoJames Vacation Home"
            )
            flash(f"ส่งคำขอยกเลิกแล้ว ยอดคืนตามนโยบายเบื้องต้น {refund_amount:.2f} บาท")
            return redirect("/my-bookings")

        conn.close()
        refund_preview = calculate_refund_amount(booking, "user")
        return render_template("cancel_request.html", booking=booking, refund_preview=refund_preview, banks=BANKS)

    conn.close()
    flash("ไม่สามารถยกเลิกรายการนี้ได้")
    return redirect("/admin" if is_admin else "/my-bookings")


@app.route("/admin/approve-cancel/<int:booking_id>")
def admin_approve_cancel_request(booking_id):
    if session.get("role") != "admin":
        return "สำหรับผู้ดูแลระบบเท่านั้น"

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM bookings WHERE id = ?", (booking_id,))
    booking = cur.fetchone()

    if booking is None:
        conn.close()
        flash("ไม่พบรายการจอง")
        return redirect("/admin")

    if booking["status"] != "cancel_requested":
        conn.close()
        flash("รายการนี้ไม่ได้อยู่ในสถานะรอจัดการการยกเลิก")
        return redirect("/admin")

    refund_amount = calculate_refund_amount(booking, booking["cancelled_by"] if booking["cancelled_by"] else "user")
    cur.execute(
        """
        UPDATE bookings
        SET status = 'refund_pending',
            refund_amount = ?,
            refund_note = COALESCE(refund_note, ?)
        WHERE id = ?
        """,
        (refund_amount, "รอผู้ดูแลระบบโอนคืนเงินให้ลูกค้า", booking_id)
    )
    conn.commit()
    conn.close()

    notify_user(
        booking["username"],
        "อนุมัติการยกเลิกแล้ว",
        f"คำขอยกเลิกการจองห้อง {booking['room_name']} ของคุณได้รับการอนุมัติแล้ว ระบบจะคืนเงิน {refund_amount:.2f} บาท ภายใน 2-4 วันทำการ"
    )
    flash("อนุมัติคำขอยกเลิกแล้ว ระบบจะรอคืนเงิน")
    return redirect("/admin")


@app.route("/admin/reject-cancel/<int:booking_id>")
def admin_reject_cancel_request(booking_id):
    if session.get("role") != "admin":
        return "สำหรับผู้ดูแลระบบเท่านั้น"

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM bookings WHERE id = ?", (booking_id,))
    booking = cur.fetchone()

    if booking is None:
        conn.close()
        flash("ไม่พบรายการจอง")
        return redirect("/admin")

    if booking["status"] != "cancel_requested":
        conn.close()
        flash("รายการนี้ไม่ได้อยู่ในสถานะรอจัดการการยกเลิก")
        return redirect("/admin")

    rollback_status = booking["cancel_requested_from"] if booking["cancel_requested_from"] else "fully_paid"
    cur.execute(
        """
        UPDATE bookings
        SET status = ?,
            cancel_requested_from = NULL,
            cancel_requested_at = NULL
        WHERE id = ?
        """,
        (rollback_status, booking_id)
    )
    conn.commit()
    conn.close()

    notify_user(
        booking["username"],
        "ไม่อนุมัติการยกเลิก",
        f"คำขอยกเลิกการจองห้อง {booking['room_name']} ของคุณไม่ได้รับการอนุมัติ รายการจองยังคงเดิม"
    )
    flash("ปฏิเสธคำขอยกเลิกแล้ว")
    return redirect("/admin")


@app.route("/admin/refund-done/<int:booking_id>")
def admin_refund_done(booking_id):
    if session.get("role") != "admin":
        return "สำหรับผู้ดูแลระบบเท่านั้น"

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM bookings WHERE id = ?", (booking_id,))
    booking = cur.fetchone()

    if booking is None:
        conn.close()
        flash("ไม่พบรายการจอง")
        return redirect("/admin")

    if booking["status"] != "refund_pending":
        conn.close()
        flash("รายการนี้ไม่ได้อยู่ในสถานะรอคืนเงิน")
        return redirect("/admin")

    now_str = get_now().strftime("%Y-%m-%d %H:%M:%S")
    cur.execute(
        """
        UPDATE bookings
        SET status = 'refunded',
            refund_completed_at = ?,
            cancel_requested_from = NULL,
            cancel_requested_at = NULL
        WHERE id = ?
        """,
        (now_str, booking_id)
    )
    conn.commit()
    conn.close()

    notify_user(
        booking["username"],
        "คืนเงินเรียบร้อยแล้ว",
        f"รายการจองห้อง {booking['room_name']} ของคุณได้รับการคืนเงินเรียบร้อยแล้ว จำนวน {float(booking['refund_amount'] or 0):.2f} บาท"
    )
    flash("อัปเดตเป็นคืนเงินแล้ว")
    return redirect("/admin")



@app.route("/admin/export-refunds")
def admin_export_refunds():
    if session.get("role") != "admin":
        return "สำหรับผู้ดูแลระบบเท่านั้น"

    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, username, room_name, paytype, paid, status, refund_method, refund_bank, refund_account,
               refund_amount, refund_note, cancelled_by, cancel_requested_at, refund_completed_at
        FROM bookings
        WHERE status IN ('cancel_requested', 'refund_pending', 'refunded')
        ORDER BY id DESC
    """)
    rows = cur.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Refunds"

    headers = [
        "Booking ID", "Username", "Room", "Payment Type", "Paid (THB)", "Refund Status",
        "Refund Amount (THB)", "Refund Method", "Bank", "Account", "Cancelled By",
        "Requested At", "Completed At", "Note"
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="5B3DB5")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="DDDDDD")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)

    for row in rows:
        bank_meta = get_bank_meta(row["refund_bank"])
        ws.append([
            row["id"],
            row["username"],
            row["room_name"],
            "จ่ายเต็ม" if row["paytype"] == "full" else "มัดจำ 25%",
            float(row["paid"] or 0),
            row["status"],
            float(row["refund_amount"] or 0),
            "PromptPay" if row["refund_method"] == "promptpay" else ("บัญชีธนาคาร" if row["refund_method"] == "bank" else row["refund_method"] or "-"),
            bank_meta["name"] if row["refund_bank"] else "-",
            row["refund_account"] or "-",
            row["cancelled_by"] or "-",
            row["cancel_requested_at"] or "-",
            row["refund_completed_at"] or "-",
            row["refund_note"] or "-"
        ])

    widths = {"A":12, "B":18, "C":18, "D":15, "E":14, "F":18, "G":18, "H":16, "I":18, "J":20, "K":14, "L":20, "M":20, "N":40}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    xlsx_data = BytesIO()
    wb.save(xlsx_data)
    xlsx_data.seek(0)
    return send_file(
        xlsx_data,
        as_attachment=True,
        download_name="refunds_gojames.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/approve/<int:booking_id>")
def approve_booking(booking_id):
    auto_cancel_expired_bookings()

    if session.get("role") != "admin":
        return "สำหรับผู้ดูแลระบบเท่านั้น"

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM bookings WHERE id = ?", (booking_id,))
    booking = cur.fetchone()

    if booking is None:
        conn.close()
        flash("ไม่พบรายการจอง")
        return redirect("/admin")

    if booking["status"] != "waiting_approval":
        conn.close()
        flash("รายการนี้ไม่ได้อยู่ในสถานะรอชำระเงิน")
        return redirect("/admin")

    room_id = booking["room_id"]
    new_in = datetime.strptime(f"{booking['checkin']} {booking['checkin_time']}", "%Y-%m-%d %H:%M")
    new_out = datetime.strptime(f"{booking['checkout']} {booking['checkout_time']}", "%Y-%m-%d %H:%M")

    placeholders = ",".join("?" for _ in BOOKED_STATUSES)
    cur.execute(f"""
        SELECT * FROM bookings
        WHERE room_id = ? AND id != ? AND status IN ({placeholders})
    """, (room_id, booking_id, *BOOKED_STATUSES))
    existing_bookings = cur.fetchall()

    for row in existing_bookings:
        exist_in = datetime.strptime(f"{row['checkin']} {row['checkin_time']}", "%Y-%m-%d %H:%M")
        exist_out = datetime.strptime(f"{row['checkout']} {row['checkout_time']}", "%Y-%m-%d %H:%M")
        if is_overlap(new_in, new_out, exist_in, exist_out):
            conn.close()
            flash("ดำเนินการไม่ได้ เพราะช่วงวันและเวลาซ้อนกับรายการที่มีอยู่แล้ว")
            return redirect("/admin")

    cur.execute("SELECT * FROM room_blocks WHERE room_id = ?", (room_id,))
    blocks = cur.fetchall()

    for block in blocks:
        block_start = datetime.strptime(block["start_date"], "%Y-%m-%d").date()
        block_end = datetime.strptime(block["end_date"], "%Y-%m-%d").date()
        if block_overlaps_booking(new_in, new_out, block_start, block_end):
            conn.close()
            flash("ดำเนินการไม่ได้ เพราะห้องนี้ถูกผู้ดูแลระบบตั้งเป็นไม่ว่างในช่วงดังกล่าว")
            return redirect("/admin")

    approved_at = get_now()
    deadline = approved_at + timedelta(minutes=15)

    cur.execute("""
        UPDATE bookings
        SET status = 'approved',
            approved_at = ?,
            payment_deadline = ?
        WHERE id = ?
    """, (
        approved_at.strftime("%Y-%m-%d %H:%M:%S"),
        deadline.strftime("%Y-%m-%d %H:%M:%S"),
        booking_id
    ))
    conn.commit()
    conn.close()

    notify_user(
        booking["username"],
        "รอชำระเงิน",
        f"รายการจองห้อง {booking['room_name']} ของคุณอยู่ในสถานะรอชำระเงิน กรุณาชำระเงินภายในเวลาที่กำหนด",
        email_subject="กรุณาชำระเงินภายในเวลาที่กำหนด",
        email_body=(
            f"รายการจองห้อง {booking['room_name']} ของคุณได้รับการอนุมัติแล้ว\n"
            f"กรุณาชำระเงินภายใน 15 นาที\n"
            f"กำหนดชำระ: {deadline.strftime('%Y-%m-%d %H:%M:%S')}"
        )
    )
    flash("รายการอยู่ในสถานะรอชำระเงินแล้ว")
    return redirect("/admin")


@app.route("/reject/<int:booking_id>")
def reject_booking(booking_id):
    if session.get("role") != "admin":
        return "สำหรับผู้ดูแลระบบเท่านั้น"

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM bookings WHERE id = ?", (booking_id,))
    booking = cur.fetchone()

    if booking is None:
        conn.close()
        flash("ไม่พบรายการจอง")
        return redirect("/admin")

    cur.execute("UPDATE bookings SET status = 'rejected' WHERE id = ?", (booking_id,))
    conn.commit()
    conn.close()

    notify_user(booking["username"], "การจองถูกปฏิเสธ", f"รายการจองห้อง {booking['room_name']} ของคุณถูกปฏิเสธ")
    flash("ปฏิเสธการจองแล้ว")
    return redirect("/admin")


@app.route("/payment/<int:booking_id>")
def payment(booking_id):
    auto_cancel_expired_bookings()

    if "user" not in session:
        flash("กรุณาเข้าสู่ระบบก่อน")
        return redirect("/login")

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM bookings WHERE id = ?", (booking_id,))
    booking = cur.fetchone()

    if booking is None:
        conn.close()
        flash("ไม่พบรายการจอง")
        return redirect("/my-bookings")

    if session.get("role") != "admin" and booking["username"] != session["user"]:
        conn.close()
        flash("คุณไม่มีสิทธิ์ดูรายการนี้")
        return redirect("/my-bookings")

    if booking["status"] != "approved":
        conn.close()
        flash("รายการนี้ยังไม่อยู่ในสถานะชำระเงิน")
        return redirect("/my-bookings")

    if booking["batch_id"]:
        cur.execute("""
            SELECT * FROM bookings
            WHERE batch_id = ? AND username = ?
            ORDER BY checkin ASC, checkin_time ASC
        """, (booking["batch_id"], booking["username"]))
        batch_bookings = cur.fetchall()
    else:
        batch_bookings = [booking]
    conn.close()

    amount = round(sum(row["paid"] for row in batch_bookings), 2)
    total = round(sum(row["total"] for row in batch_bookings), 2)
    remaining_due = round(sum(row["remaining_due"] for row in batch_bookings), 2)
    qr_filename = create_qr(booking["id"], amount)

    return render_template(
        "payment.html",
        booking=booking,
        booking_id=booking["id"],
        batch_bookings=batch_bookings,
        amount=amount,
        remaining_due=remaining_due,
        qr_filename=qr_filename,
        paytype=booking["paytype"],
        total=total,
        payment_mode="initial"
    )



@app.route("/remaining-payment/<int:booking_id>")
def remaining_payment(booking_id):
    auto_cancel_expired_bookings()

    if "user" not in session:
        flash("กรุณาเข้าสู่ระบบก่อน")
        return redirect("/login")

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM bookings WHERE id = ?", (booking_id,))
    booking = cur.fetchone()

    if booking is None:
        conn.close()
        flash("ไม่พบรายการจอง")
        return redirect("/my-bookings")

    if booking["username"] != session["user"]:
        conn.close()
        flash("คุณไม่มีสิทธิ์ดูรายการนี้")
        return redirect("/my-bookings")

    allowed_remaining_statuses = {"deposit_paid", "remaining_counter_pending"}
    if booking["status"] not in allowed_remaining_statuses:
        conn.close()
        flash("รายการนี้ยังไม่อยู่ในสถานะชำระส่วนที่เหลือ")
        return redirect("/my-bookings")

    if booking["batch_id"]:
        cur.execute("""
            SELECT * FROM bookings
            WHERE batch_id = ? AND username = ?
            ORDER BY checkin ASC, checkin_time ASC
        """, (booking["batch_id"], booking["username"]))
        batch_bookings = [row for row in cur.fetchall() if row["status"] in allowed_remaining_statuses]
    else:
        batch_bookings = [booking]
    conn.close()

    amount = round(sum(row["remaining_due"] for row in batch_bookings), 2)
    total = round(sum(row["total"] for row in batch_bookings), 2)
    paid_so_far = round(sum(row["paid"] for row in batch_bookings), 2)
    qr_filename = create_qr(booking["id"], amount)

    return render_template(
        "payment.html",
        booking=booking,
        booking_id=booking["id"],
        batch_bookings=batch_bookings,
        amount=amount,
        paid_so_far=paid_so_far,
        remaining_due=0,
        qr_filename=qr_filename,
        paytype=booking["paytype"],
        total=total,
        payment_mode="remaining"
    )


@app.route("/upload_slip/<int:booking_id>", methods=["POST"])
def upload_slip(booking_id):
    auto_cancel_expired_bookings()

    if "user" not in session:
        flash("กรุณาเข้าสู่ระบบก่อน")
        return redirect("/login")

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM bookings WHERE id = ?", (booking_id,))
    booking = cur.fetchone()

    if booking is None:
        conn.close()
        flash("ไม่พบรายการจอง")
        return redirect("/my-bookings")

    if booking["username"] != session["user"]:
        conn.close()
        flash("คุณไม่มีสิทธิ์อัปโหลดสลิปของรายการนี้")
        return redirect("/my-bookings")

    valid_statuses = {"approved", "deposit_paid", "remaining_counter_pending"}
    if booking["status"] not in valid_statuses:
        conn.close()
        flash("รายการนี้ยังไม่สามารถอัปโหลดสลิปได้")
        return redirect("/my-bookings")

    is_remaining_flow = booking["status"] in {"deposit_paid", "remaining_counter_pending"}
    remain_method = request.form.get("remain_method", "qr") if is_remaining_flow else "qr"
    payment_method = request.form.get("payment_method", "promptpay") if not is_remaining_flow else remain_method

    # กรณีลูกค้าเลือกชำระทั้งหมดหน้าเคาน์เตอร์ตั้งแต่แรก
    if (not is_remaining_flow) and payment_method == "counter":
        if booking["batch_id"]:
            cur.execute("""
                UPDATE bookings
                SET status = 'counter_pending', payment_method = 'counter'
                WHERE batch_id = ? AND username = ? AND status = 'approved'
            """, (booking["batch_id"], booking["username"]))
        else:
            cur.execute("""
                UPDATE bookings
                SET status = 'counter_pending', payment_method = 'counter'
                WHERE id = ?
            """, (booking_id,))
        conn.commit()
        conn.close()
        flash("บันทึกแล้ว: รายการนี้จะชำระเงินหน้าเคาน์เตอร์")
        notify_user(booking["username"], "เลือกชำระหน้าเคาน์เตอร์", f"รายการจองห้อง {booking['room_name']} จะชำระเงินที่หน้าเคาน์เตอร์")
        notify_admin(
            "ลูกค้าเลือกชำระหน้าเคาน์เตอร์",
            f"ลูกค้า {booking['username']} เลือกชำระเงินหน้าเคาน์เตอร์สำหรับห้อง {booking['room_name']}"
        )
        return redirect("/my-bookings")

    if is_remaining_flow and remain_method == "counter":
        if booking["batch_id"]:
            cur.execute("""
                UPDATE bookings
                SET status = 'remaining_counter_pending', remaining_payment_method = 'counter'
                WHERE batch_id = ? AND username = ? AND status IN ('deposit_paid', 'remaining_counter_pending')
            """, (booking["batch_id"], booking["username"]))
        else:
            cur.execute("""
                UPDATE bookings
                SET status = 'remaining_counter_pending', remaining_payment_method = 'counter'
                WHERE id = ?
            """, (booking_id,))
        conn.commit()
        conn.close()
        flash("บันทึกแล้ว: ยอดส่วนที่เหลือจะชำระหน้าเคาน์เตอร์ที่พัก")
        notify_user(booking["username"], "เลือกชำระยอดคงเหลือหน้าเคาน์เตอร์", f"ยอดคงเหลือของรายการจองห้อง {booking['room_name']} จะชำระหน้าเคาน์เตอร์")
        notify_admin(
            "ลูกค้าเลือกชำระยอดคงเหลือหน้าเคาน์เตอร์",
            f"ลูกค้า {booking['username']} เลือกชำระยอดคงเหลือหน้าเคาน์เตอร์สำหรับห้อง {booking['room_name']}"
        )
        return redirect("/my-bookings")

    slip = request.files.get("slip")
    if not slip or slip.filename == "":
        conn.close()
        flash("กรุณาเลือกไฟล์สลิป")
        return redirect(url_for("payment", booking_id=booking_id) if booking["status"] == "approved" else url_for("remaining_payment", booking_id=booking_id))

    if not allowed_file(slip.filename):
        conn.close()
        flash("รองรับเฉพาะไฟล์ png, jpg, jpeg, webp")
        return redirect(url_for("payment", booking_id=booking_id) if booking["status"] == "approved" else url_for("remaining_payment", booking_id=booking_id))

    ext = slip.filename.rsplit(".", 1)[1].lower()
    filename = secure_filename(f"slip_{booking_id}.{ext}")
    filepath = os.path.join(UPLOAD_DIR, filename)
    slip.save(filepath)

    target_status = "waiting_slip_approval" if booking["status"] == "approved" else "waiting_remaining_slip_approval"

    if booking["batch_id"]:
        if is_remaining_flow:
            cur.execute(f"""
                UPDATE bookings
                SET slip_filename = ?, status = ?, remaining_payment_method = 'qr', payment_method = 'promptpay'
                WHERE batch_id = ? AND username = ? AND status IN ('deposit_paid', 'remaining_counter_pending')
            """, (filename, target_status, booking["batch_id"], booking["username"]))
        else:
            cur.execute(f"""
                UPDATE bookings
                SET slip_filename = ?, status = ?, payment_method = ?
                WHERE batch_id = ? AND username = ? AND status = ?
            """, (filename, target_status, payment_method, booking["batch_id"], booking["username"], booking["status"]))
    else:
        if is_remaining_flow:
            cur.execute("""
                UPDATE bookings
                SET slip_filename = ?, status = ?, remaining_payment_method = 'qr', payment_method = 'promptpay'
                WHERE id = ?
            """, (filename, target_status, booking_id))
        else:
            cur.execute("""
                UPDATE bookings
                SET slip_filename = ?, status = ?, payment_method = ?
                WHERE id = ?
            """, (filename, target_status, payment_method, booking_id))
    conn.commit()
    conn.close()

    notify_user(
        booking["username"],
        "อัปโหลดหลักฐานการชำระเงินแล้ว",
        f"ระบบได้รับสลิปของรายการจองห้อง {booking['room_name']} แล้ว กรุณารอผู้ดูแลระบบตรวจสอบ"
    )
    notify_admin(
        "มีหลักฐานการชำระเงินรอตรวจสอบ",
        f"ลูกค้า {booking['username']} อัปโหลดหลักฐานการชำระเงินสำหรับห้อง {booking['room_name']} กรุณาตรวจสอบ",
        email_subject="มีหลักฐานการชำระเงินรอตรวจสอบ - GoJames Vacation Home",
        email_body=(
            f"ลูกค้า {booking['username']} อัปโหลดหลักฐานการชำระเงินใหม่\n"
            f"ห้องพัก: {booking['room_name']}\n"
            "กรุณาเข้าสู่ระบบผู้ดูแลระบบเพื่อตรวจสอบการชำระเงิน"
        )
    )
    flash("อัปโหลดหลักฐานการชำระเงินเรียบร้อยแล้ว กรุณารอผู้ดูแลระบบตรวจสอบ")
    return redirect("/my-bookings")

@app.route("/approve_slip/<int:booking_id>")
def approve_slip(booking_id):
    if session.get("role") != "admin":
        return "สำหรับผู้ดูแลระบบเท่านั้น"

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM bookings WHERE id = ?", (booking_id,))
    booking = cur.fetchone()

    if booking is None:
        conn.close()
        flash("ไม่พบรายการจอง")
        return redirect("/admin")

    if booking["status"] not in ("waiting_slip_approval", "waiting_remaining_slip_approval"):
        conn.close()
        flash("รายการนี้ไม่ได้อยู่ในสถานะรอตรวจสอบการชำระเงิน")
        return redirect("/admin")

    if booking["status"] == "waiting_slip_approval":
        new_status = "fully_paid" if booking["paytype"] == "full" else "deposit_paid"
        update_paid_sql = ""
        params = ()
        notify_msg = "รายการจองห้อง {room} ของคุณได้รับการยืนยันการชำระเงินเรียบร้อยแล้ว"
    else:
        new_status = "fully_paid"
        notify_msg = "รายการจองห้อง {room} ของคุณได้รับการยืนยันการชำระส่วนที่เหลือเรียบร้อยแล้ว"

    if booking["batch_id"]:
        if booking["status"] == "waiting_remaining_slip_approval":
            cur.execute("""
                UPDATE bookings
                SET status = ?, paid = total, remaining_due = 0,
                    receipt_no = COALESCE(receipt_no, ?),
                    receipt_issued_at = COALESCE(receipt_issued_at, ?)
                WHERE batch_id = ? AND username = ? AND status = 'waiting_remaining_slip_approval'
            """, (new_status, make_receipt_no(booking_id), db_now_str(), booking["batch_id"], booking["username"]))
        else:
            if new_status == "fully_paid":
                cur.execute("""
                    UPDATE bookings
                    SET status = ?,
                        receipt_no = COALESCE(receipt_no, ?),
                        receipt_issued_at = COALESCE(receipt_issued_at, ?)
                    WHERE batch_id = ? AND username = ? AND status = 'waiting_slip_approval'
                """, (new_status, make_receipt_no(booking_id), db_now_str(), booking["batch_id"], booking["username"]))
            else:
                cur.execute("""
                    UPDATE bookings
                    SET status = ?
                    WHERE batch_id = ? AND username = ? AND status = 'waiting_slip_approval'
                """, (new_status, booking["batch_id"], booking["username"]))
    else:
        if booking["status"] == "waiting_remaining_slip_approval":
            cur.execute("UPDATE bookings SET status = ?, paid = total, remaining_due = 0, receipt_no = COALESCE(receipt_no, ?), receipt_issued_at = COALESCE(receipt_issued_at, ?) WHERE id = ?", (new_status, make_receipt_no(booking_id), db_now_str(), booking_id))
        else:
            if new_status == "fully_paid":
                cur.execute("UPDATE bookings SET status = ?, receipt_no = COALESCE(receipt_no, ?), receipt_issued_at = COALESCE(receipt_issued_at, ?) WHERE id = ?", (new_status, make_receipt_no(booking_id), db_now_str(), booking_id))
            else:
                cur.execute("UPDATE bookings SET status = ? WHERE id = ?", (new_status, booking_id))

    conn.commit()
    conn.close()

    notify_user(booking["username"], "ชำระเงินสำเร็จ", notify_msg.format(room=booking['room_name']))
    flash("ยืนยันการชำระเงินเรียบร้อยแล้ว")
    return redirect("/admin")


@app.route("/reject_slip/<int:booking_id>")
def reject_slip(booking_id):
    if session.get("role") != "admin":
        return "สำหรับผู้ดูแลระบบเท่านั้น"

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM bookings WHERE id = ?", (booking_id,))
    booking = cur.fetchone()

    if booking is None:
        conn.close()
        flash("ไม่พบรายการจอง")
        return redirect("/admin")

    if booking["status"] not in ("waiting_slip_approval", "waiting_remaining_slip_approval"):
        conn.close()
        flash("รายการนี้ไม่ได้อยู่ในสถานะรอตรวจสอบการชำระเงิน")
        return redirect("/admin")

    rollback_status = "approved" if booking["status"] == "waiting_slip_approval" else "deposit_paid"

    if booking["batch_id"]:
        cur.execute("""
            UPDATE bookings
            SET status = ?, slip_filename = NULL
            WHERE batch_id = ? AND username = ? AND status = ?
        """, (rollback_status, booking["batch_id"], booking["username"], booking["status"]))
    else:
        cur.execute("""
            UPDATE bookings
            SET status = ?, slip_filename = NULL
            WHERE id = ?
        """, (rollback_status, booking_id))
    conn.commit()
    conn.close()

    msg = "หลักฐานการชำระเงิน" if rollback_status == "approved" else "หลักฐานการชำระเงินส่วนที่เหลือ"
    notify_user(booking["username"], "หลักฐานการชำระเงินไม่ถูกต้อง", f"{msg} สำหรับห้อง {booking['room_name']} ของคุณถูกปฏิเสธ กรุณาอัปโหลดใหม่")
    flash("ส่งกลับให้แก้ไขแล้ว ให้ลูกค้าอัปโหลดใหม่")
    return redirect("/admin")


@app.route("/block-room", methods=["POST"])
def block_room():
    if session.get("role") != "admin":
        return "สำหรับผู้ดูแลระบบเท่านั้น"

    room_id = request.form["room_id"]
    start_date = request.form["start_date"]
    end_date = request.form["end_date"]
    note = request.form.get("note", "").strip()

    try:
        room_id = int(room_id)
        start_day = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_day = datetime.strptime(end_date, "%Y-%m-%d").date()
    except ValueError:
        flash("ข้อมูลวันที่ไม่ถูกต้อง")
        return redirect("/admin")

    if end_day < start_day:
        flash("วันสิ้นสุดต้องไม่ก่อนวันเริ่มต้น")
        return redirect("/admin")

    conn = get_db()
    cur = conn.cursor()

    placeholders = ",".join("?" for _ in BOOKED_STATUSES)
    cur.execute(f"""
        SELECT * FROM bookings
        WHERE room_id = ? AND status IN ({placeholders})
    """, (room_id, *BOOKED_STATUSES))
    bookings = cur.fetchall()

    for booking in bookings:
        checkin = datetime.strptime(f"{booking['checkin']} {booking['checkin_time']}", "%Y-%m-%d %H:%M")
        checkout = datetime.strptime(f"{booking['checkout']} {booking['checkout_time']}", "%Y-%m-%d %H:%M")
        if block_overlaps_booking(checkin, checkout, start_day, end_day):
            conn.close()
            flash("บล็อกไม่ได้ เพราะมีรายการจองหรือชำระแล้วในช่วงดังกล่าว")
            return redirect("/admin")

    cur.execute("""
        INSERT INTO room_blocks(room_id, start_date, end_date, note, created_at, status)
        VALUES(?,?,?,?,?,?)
    """, (room_id, start_date, end_date, note, get_now().strftime("%Y-%m-%d %H:%M:%S"), "blocked"))
    conn.commit()
    conn.close()

    notify_admin(
        "บล็อกห้องเรียบร้อยแล้ว",
        f"ห้อง {room_id} ถูกตั้งค่าไม่ว่างตั้งแต่ {start_date} ถึง {end_date}" + (f" หมายเหตุ: {note}" if note else "")
    )
    flash("บล็อกห้องเรียบร้อยแล้ว")
    return redirect("/admin")


@app.route("/unblock-room/<int:block_id>")
def unblock_room(block_id):
    if session.get("role") != "admin":
        return "สำหรับผู้ดูแลระบบเท่านั้น"

    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM room_blocks WHERE id = ?", (block_id,))
    conn.commit()
    conn.close()

    flash("ยกเลิกการบล็อกห้องแล้ว")
    return redirect("/admin")


@app.route("/notifications")
def notifications():
    if "user" not in session:
        flash("กรุณาเข้าสู่ระบบก่อน")
        return redirect("/login")

    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT * FROM notifications
        WHERE username = ?
        ORDER BY id DESC
    """, (session["user"],))
    items = cur.fetchall()
    conn.close()

    return render_template("notifications.html", notifications=items)


@app.route("/notifications/read/<int:notification_id>")
def read_notification(notification_id):
    if "user" not in session:
        flash("กรุณาเข้าสู่ระบบก่อน")
        return redirect("/login")

    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        UPDATE notifications
        SET is_read = 1
        WHERE id = ? AND username = ?
    """, (notification_id, session["user"]))
    conn.commit()
    conn.close()

    return redirect("/notifications")


@app.route("/notifications/read-all")
def read_all_notifications():
    if "user" not in session:
        flash("กรุณาเข้าสู่ระบบก่อน")
        return redirect("/login")

    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        UPDATE notifications
        SET is_read = 1
        WHERE username = ?
    """, (session["user"],))
    conn.commit()
    conn.close()

    return redirect("/notifications")


@app.route("/admin/mark-counter-paid/<int:booking_id>")
def admin_mark_counter_paid(booking_id):
    if session.get("role") != "admin":
        return "สำหรับผู้ดูแลระบบเท่านั้น"

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM bookings WHERE id = ?", (booking_id,))
    booking = cur.fetchone()

    if booking is None:
        conn.close()
        flash("ไม่พบรายการจอง")
        return redirect("/admin")

    if booking["status"] not in ("remaining_counter_pending", "counter_pending"):
        conn.close()
        flash("รายการนี้ยังไม่ได้เลือกชำระหน้าเคาน์เตอร์")
        return redirect("/admin")

    if booking["batch_id"]:
        cur.execute("""
            UPDATE bookings
            SET status = 'fully_paid', paid = total, remaining_due = 0,
                payment_method = COALESCE(payment_method, 'counter'),
                receipt_no = COALESCE(receipt_no, ?),
                receipt_issued_at = COALESCE(receipt_issued_at, ?)
            WHERE batch_id = ? AND username = ? AND status IN ('remaining_counter_pending', 'counter_pending')
        """, (make_receipt_no(booking_id), db_now_str(), booking["batch_id"], booking["username"]))
    else:
        cur.execute("""
            UPDATE bookings
            SET status = 'fully_paid', paid = total, remaining_due = 0,
                payment_method = COALESCE(payment_method, 'counter'),
                receipt_no = COALESCE(receipt_no, ?),
                receipt_issued_at = COALESCE(receipt_issued_at, ?)
            WHERE id = ?
        """, (make_receipt_no(booking_id), db_now_str(), booking_id))

    conn.commit()
    conn.close()

    notify_user(booking["username"], "ชำระเงินครบแล้ว", f"รายการจองห้อง {booking['room_name']} ของคุณได้รับการยืนยันว่าชำระยอดคงเหลือหน้าเคาน์เตอร์เรียบร้อยแล้ว")
    flash("บันทึกการชำระหน้าเคาน์เตอร์เรียบร้อยแล้ว")
    return redirect("/admin")



@app.route("/admin/check-in/<int:booking_id>")
def admin_check_in(booking_id):
    if session.get("role") != "admin":
        return "สำหรับผู้ดูแลระบบเท่านั้น"

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM bookings WHERE id = ?", (booking_id,))
    booking = cur.fetchone()

    if booking is None:
        conn.close()
        flash("ไม่พบรายการจอง")
        return redirect("/admin")

    if booking["status"] != "fully_paid":
        conn.close()
        flash("เช็กอินปกติได้เฉพาะรายการที่ชำระครบแล้ว หากชำระมัดจำให้ใช้ปุ่มรับยอดคงเหลือหน้าเคาน์เตอร์และเช็กอิน")
        return redirect("/admin")

    # ผู้ดูแลระบบสามารถกดเข้าพักก่อนเวลาได้ เช่น ลูกค้ามาถึงก่อนเวลาเช็กอิน
    now_str = db_now_str()

    if booking["batch_id"]:
        cur.execute("""
            UPDATE bookings
            SET status = 'checked_in',
                checked_in_at = ?
            WHERE batch_id = ? AND username = ? AND status = 'fully_paid'
        """, (now_str, booking["batch_id"], booking["username"]))
    else:
        cur.execute("""
            UPDATE bookings
            SET status = 'checked_in',
                checked_in_at = ?
            WHERE id = ?
        """, (now_str, booking_id))

    conn.commit()
    conn.close()

    notify_user(
        booking["username"],
        "เข้าพักแล้ว",
        f"รายการจองห้อง {booking['room_name']} ได้ถูกเปลี่ยนสถานะเป็นเข้าพักแล้ว"
    )
    flash("เปลี่ยนสถานะเป็นเข้าพักแล้วเรียบร้อย")
    return redirect("/admin")

@app.route("/admin/counter-check-in/<int:booking_id>", methods=["POST"])
def admin_counter_check_in(booking_id):
    """ผู้ดูแลระบบรับยอดคงเหลือหน้าเคาน์เตอร์แล้วเช็กอินทันที"""
    if session.get("role") != "admin":
        return "สำหรับผู้ดูแลระบบเท่านั้น"

    pay_method = request.form.get("counter_method", "cash")
    pay_method_text = "เงินสด" if pay_method == "cash" else "โอน"

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM bookings WHERE id = ?", (booking_id,))
    booking = cur.fetchone()

    if booking is None:
        conn.close()
        flash("ไม่พบรายการจอง")
        return redirect("/admin")

    if booking["status"] not in ("deposit_paid", "counter_pending"):
        conn.close()
        flash("รายการนี้ไม่ได้อยู่ในสถานะที่สามารถรับเงินหน้าเคาน์เตอร์และเช็กอินได้")
        return redirect("/admin")

    # ผู้ดูแลระบบสามารถรับเงินหน้าเคาน์เตอร์และเช็กอินก่อนเวลาได้
    now_str = db_now_str()
    note_suffix = f" | ชำระยอดคงเหลือหน้าเคาน์เตอร์ด้วย{pay_method_text} และเช็กอินแล้ว {now_str}"

    if booking["batch_id"]:
        cur.execute("""
            UPDATE bookings
            SET status = 'checked_in',
                paid = total,
                remaining_due = 0,
                remaining_payment_method = ?,
                payment_method = ?,
                checked_in_at = ?,
                receipt_no = COALESCE(receipt_no, ?),
                receipt_issued_at = COALESCE(receipt_issued_at, ?),
                note = COALESCE(note, '') || ?
            WHERE batch_id = ? AND username = ? AND status IN ('deposit_paid','counter_pending')
        """, (pay_method, pay_method, now_str, make_receipt_no(booking_id), db_now_str(), note_suffix, booking["batch_id"], booking["username"]))
    else:
        cur.execute("""
            UPDATE bookings
            SET status = 'checked_in',
                paid = total,
                remaining_due = 0,
                remaining_payment_method = ?,
                payment_method = ?,
                checked_in_at = ?,
                receipt_no = COALESCE(receipt_no, ?),
                receipt_issued_at = COALESCE(receipt_issued_at, ?),
                note = COALESCE(note, '') || ?
            WHERE id = ?
        """, (pay_method, pay_method, now_str, make_receipt_no(booking_id), db_now_str(), note_suffix, booking_id))

    conn.commit()
    conn.close()

    notify_user(
        booking["username"],
        "เข้าพักแล้ว",
        f"รายการจองห้อง {booking['room_name']} ได้ชำระยอดคงเหลือหน้าเคาน์เตอร์และเช็กอินเรียบร้อยแล้ว"
    )
    flash(f"บันทึกชำระยอดคงเหลือด้วย{pay_method_text} และเช็กอินเรียบร้อยแล้ว")
    return redirect("/admin")


@app.route("/admin/no-show/<int:booking_id>", methods=["POST"])
def admin_no_show(booking_id):
    """กรณีลูกค้าจอง/จ่ายแล้ว แต่ไม่มาเมื่อเลยเวลาตัดสิทธิ์ ให้ปล่อยห้องกลับมาว่าง"""
    if session.get("role") != "admin":
        return "สำหรับผู้ดูแลระบบเท่านั้น"

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM bookings WHERE id = ?", (booking_id,))
    booking = cur.fetchone()

    if booking is None:
        conn.close()
        flash("ไม่พบรายการจอง")
        return redirect("/admin")

    allowed_statuses = ("approved", "deposit_paid", "fully_paid", "remaining_counter_pending", "counter_pending")
    if booking["status"] not in allowed_statuses:
        conn.close()
        flash("สถานะนี้ไม่สามารถเปลี่ยนเป็นลูกค้าไม่มาได้")
        return redirect("/admin")

    cutoff_dt = get_no_show_cutoff_dt(booking)
    if cutoff_dt and get_now() < cutoff_dt:
        conn.close()
        flash(f"ยังไม่ถึงเวลาตัดสิทธิ์ No Show ({NO_SHOW_CUTOFF_TIME} น.) จึงยังไม่ควรปล่อยห้อง")
        return redirect("/admin")

    now_str = db_now_str()
    note_suffix = f" | No Show: ลูกค้าไม่มาเกินเวลาตัดสิทธิ์ {NO_SHOW_CUTOFF_TIME} น. ผู้ดูแลระบบปล่อยห้องกลับมาว่าง {now_str}"

    if booking["batch_id"]:
        cur.execute("""
            UPDATE bookings
            SET status = 'no_show',
                note = COALESCE(note, '') || ?
            WHERE batch_id = ? AND username = ? AND status IN ('approved','deposit_paid','fully_paid','remaining_counter_pending','counter_pending')
        """, (note_suffix, booking["batch_id"], booking["username"]))
    else:
        cur.execute("""
            UPDATE bookings
            SET status = 'no_show',
                note = COALESCE(note, '') || ?
            WHERE id = ?
        """, (note_suffix, booking_id))

    conn.commit()
    conn.close()

    notify_user(
        booking["username"],
        "รายการจองถูกปรับเป็น No Show",
        f"รายการจองห้อง {booking['room_name']} ถูกปรับเป็น No Show เนื่องจากเลยเวลาตัดสิทธิ์และยังไม่มาเข้าพัก"
    )
    flash("เปลี่ยนสถานะเป็น No Show แล้ว ห้องถูกปล่อยกลับมาว่างและสามารถรับ Walk-in ได้")
    return redirect("/admin")

@app.route("/admin/check-out/<int:booking_id>", methods=["POST"])
def admin_check_out(booking_id):
    if session.get("role") != "admin":
        return "สำหรับผู้ดูแลระบบเท่านั้น"

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM bookings WHERE id = ?", (booking_id,))
    booking = cur.fetchone()

    if booking is None:
        conn.close()
        flash("ไม่พบรายการจอง")
        return redirect("/admin")

    if booking["status"] != "checked_in":
        conn.close()
        flash("เช็กเอาท์ได้เฉพาะรายการที่อยู่ในสถานะเข้าพักแล้ว")
        return redirect("/admin")

    now_str = db_now_str()
    if booking["batch_id"]:
        cur.execute("""
            UPDATE bookings
            SET status = 'checked_out',
                checked_out_at = ?
            WHERE batch_id = ? AND username = ? AND status = 'checked_in'
        """, (now_str, booking["batch_id"], booking["username"]))
    else:
        cur.execute("""
            UPDATE bookings
            SET status = 'checked_out',
                checked_out_at = ?
            WHERE id = ?
        """, (now_str, booking_id))

    conn.commit()
    conn.close()

    notify_user(
        booking["username"],
        "เช็กเอาท์แล้ว",
        f"รายการจองห้อง {booking['room_name']} ถูกเปลี่ยนสถานะเป็นเช็กเอาท์แล้ว"
    )
    flash("เช็กเอาท์เรียบร้อยแล้ว ห้องจะกลับไปว่างสำหรับวันถัดไปตามปฏิทิน")
    return redirect("/admin")


@app.route("/admin/walk-in", methods=["POST"])
def admin_walk_in():
    if session.get("role") != "admin":
        return "สำหรับผู้ดูแลระบบเท่านั้น"

    try:
        room_id = int(request.form.get("room_id", ""))
        checkin = request.form.get("checkin", "").strip()
        checkout = request.form.get("checkout", "").strip()
        checkin_time = request.form.get("checkin_time", "14:00").strip()
        checkout_time = RESORT_CHECKOUT_GUIDE if "RESORT_CHECKOUT_GUIDE" in globals() else "12:00"
        total = float(request.form.get("total", "0") or 0)
    except ValueError:
        flash("ข้อมูล Walk-in ไม่ถูกต้อง")
        return redirect("/admin")

    guest_name = request.form.get("guest_name", "").strip() or "ลูกค้า Walk-in"
    guest_phone = request.form.get("guest_phone", "").strip()
    if guest_phone and (not guest_phone.isdigit() or len(guest_phone) != 10 or not guest_phone.startswith("0")):
        flash("กรุณากรอกเบอร์โทรศัพท์ Walk-in เป็นตัวเลข 10 หลัก และขึ้นต้นด้วย 0")
        return redirect("/admin")
    guest_phone = guest_phone or "walkin"
    pay_method = request.form.get("pay_method", "cash")
    room = get_room(room_id)

    if room is None:
        flash("ไม่พบห้องพัก")
        return redirect("/admin")

    try:
        dt1 = datetime.strptime(f"{checkin} {checkin_time}", "%Y-%m-%d %H:%M")
        dt2 = datetime.strptime(f"{checkout} {checkout_time}", "%Y-%m-%d %H:%M")
    except ValueError:
        flash("รูปแบบวันที่หรือเวลาไม่ถูกต้อง")
        return redirect("/admin")

    if dt2 <= dt1:
        flash("วัน/เวลาออกต้องมากกว่าวัน/เวลาเข้า")
        return redirect("/admin")

    if total <= 0:
        total = calculate_charge_nights(dt1, dt2) * room["price"]

    # กันจองทับทั้งรายการจองเดิมและช่วงที่บล็อกห้องไว้
    conn = get_db()
    cur = conn.cursor()
    placeholders = ",".join("?" for _ in BOOKED_STATUSES)
    cur.execute(f"""
        SELECT * FROM bookings
        WHERE room_id = ? AND status IN ({placeholders})
    """, (room_id, *BOOKED_STATUSES))
    existing = cur.fetchall()

    cur.execute("""
        SELECT * FROM room_blocks
        WHERE room_id = ?
          AND COALESCE(status, 'blocked') != 'free'
    """, (room_id,))
    blocks = cur.fetchall()
    conn.close()

    for b in existing:
        b_start = datetime.strptime(f"{b['checkin']} {b['checkin_time']}", "%Y-%m-%d %H:%M")
        b_end = datetime.strptime(f"{b['checkout']} {b['checkout_time']}", "%Y-%m-%d %H:%M")
        if is_overlap(dt1, dt2, b_start, b_end):
            flash("ช่วงเวลานี้มีรายการจอง/เข้าพักอยู่แล้ว ไม่สามารถสร้าง Walk-in ซ้ำได้")
            return redirect("/admin")

    for block in blocks:
        try:
            block_start = datetime.strptime(block["start_date"], "%Y-%m-%d").date()
            block_end = datetime.strptime(block["end_date"], "%Y-%m-%d").date()
        except Exception:
            continue
        if block_overlaps_booking(dt1, dt2, block_start, block_end):
            flash("ช่วงเวลานี้ห้องถูกบล็อกไว้ ไม่สามารถสร้าง Walk-in ได้")
            return redirect("/admin")

    username = ensure_walkin_user(guest_phone, guest_name)
    nights = calculate_charge_nights(dt1, dt2)
    now_str = db_now_str()
    method_text = "เงินสด" if pay_method == "cash" else "โอน"

    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO bookings(
            username, room_id, room_name, checkin, checkout,
            checkin_time, checkout_time, nights, total, paytype,
            paid, status, created_at, approved_at, remaining_due,
            remaining_payment_method, counter_payment_method, checked_in_at,
            payment_method, receipt_no, receipt_issued_at
        )
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        username, room_id, room["name"], checkin, checkout,
        checkin_time, checkout_time, nights, total, "full",
        total, "checked_in", now_str, now_str, 0,
        pay_method, pay_method, now_str,
        pay_method, make_receipt_no(room_id), now_str
    ))
    conn.commit()
    conn.close()

    notify_user(username, "สร้างรายการ Walk-in", f"สร้างรายการ Walk-in ห้อง {room['name']} และเช็กอินแล้ว ชำระด้วย{method_text}")
    flash(f"สร้าง Walk-in และเช็กอินห้อง {room['name']} แล้ว ชำระด้วย{method_text}")
    return redirect("/admin")


@app.route("/receipt/<int:booking_id>")
def receipt(booking_id):
    if "user" not in session:
        flash("กรุณาเข้าสู่ระบบก่อน")
        return redirect("/login")

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM bookings WHERE id = ?", (booking_id,))
    booking = cur.fetchone()

    if booking is None:
        conn.close()
        flash("ไม่พบรายการจอง")
        return redirect("/my-bookings")

    if session.get("role") != "admin" and booking["username"] != session["user"]:
        conn.close()
        flash("คุณไม่มีสิทธิ์ดูใบเสร็จนี้")
        return redirect("/my-bookings")

    if booking["status"] not in ("fully_paid", "checked_in", "checked_out"):
        conn.close()
        flash("ใบเสร็จจะแสดงได้หลังชำระครบแล้ว")
        return redirect("/my-bookings")

    if not booking["receipt_no"]:
        receipt_no = make_receipt_no(booking_id)
        issued_at = db_now_str()
        cur.execute("UPDATE bookings SET receipt_no = ?, receipt_issued_at = ? WHERE id = ?", (receipt_no, issued_at, booking_id))
        conn.commit()
        cur.execute("SELECT * FROM bookings WHERE id = ?", (booking_id,))
        booking = cur.fetchone()

    conn.close()
    return render_template("receipt.html", booking=booking)


def get_admin_report_summary():
    """สรุปข้อมูลสำหรับหน้าเครื่องมือเพิ่มเติมของผู้ดูแลระบบ"""
    conn = get_db()
    cur = conn.cursor()
    today = get_now().strftime("%Y-%m-%d")
    month_prefix = get_now().strftime("%Y-%m")

    active_revenue_statuses = (
        "deposit_paid", "remaining_counter_pending", "counter_pending",
        "waiting_remaining_slip_approval", "fully_paid", "checked_in", "checked_out",
        "refund_pending", "refunded"
    )
    placeholders = ",".join(["?"] * len(active_revenue_statuses))

    def scalar(sql, params=()):
        cur.execute(sql, params)
        row = cur.fetchone()
        return list(row)[0] if row else 0

    total_bookings = scalar("SELECT COUNT(*) FROM bookings") or 0
    pending_bookings = scalar("SELECT COUNT(*) FROM bookings WHERE status IN ('approved','waiting_slip_approval','waiting_remaining_slip_approval')") or 0
    checked_in_count = scalar("SELECT COUNT(*) FROM bookings WHERE status = 'checked_in'") or 0
    checked_out_count = scalar("SELECT COUNT(*) FROM bookings WHERE status = 'checked_out'") or 0
    cancelled_count = scalar("SELECT COUNT(*) FROM bookings WHERE status IN ('cancelled','rejected','no_show')") or 0
    walkin_count = scalar("SELECT COUNT(*) FROM bookings WHERE username LIKE 'walkin_%'") or 0
    customer_count = scalar("SELECT COUNT(DISTINCT username) FROM bookings WHERE username NOT LIKE 'walkin_%'") or 0
    blocked_count = scalar("SELECT COUNT(*) FROM room_blocks WHERE status = 'blocked'") or 0

    total_revenue = scalar(f"SELECT COALESCE(SUM(paid), 0) FROM bookings WHERE status IN ({placeholders})", active_revenue_statuses) or 0
    today_revenue = scalar(
        f"SELECT COALESCE(SUM(paid), 0) FROM bookings WHERE status IN ({placeholders}) AND substr(COALESCE(created_at,''),1,10) = ?",
        (*active_revenue_statuses, today)
    ) or 0
    month_revenue = scalar(
        f"SELECT COALESCE(SUM(paid), 0) FROM bookings WHERE status IN ({placeholders}) AND substr(COALESCE(created_at,''),1,7) = ?",
        (*active_revenue_statuses, month_prefix)
    ) or 0

    open_rooms = sum(1 for r in get_rooms_with_status() if r.get("is_open", True))
    total_rooms = len(get_rooms_with_status())
    available_today = max(0, open_rooms - checked_in_count)

    cur.execute("""
        SELECT id, room_name, username, status, total, paid, created_at
        FROM bookings
        ORDER BY id DESC
        LIMIT 5
    """)
    latest_bookings = cur.fetchall()
    conn.close()

    return {
        "today_revenue": float(today_revenue),
        "month_revenue": float(month_revenue),
        "total_revenue": float(total_revenue),
        "total_bookings": int(total_bookings),
        "pending_bookings": int(pending_bookings),
        "checked_in_count": int(checked_in_count),
        "checked_out_count": int(checked_out_count),
        "cancelled_count": int(cancelled_count),
        "walkin_count": int(walkin_count),
        "customer_count": int(customer_count),
        "total_rooms": int(total_rooms),
        "open_rooms": int(open_rooms),
        "available_today": int(available_today),
        "blocked_count": int(blocked_count),
        "latest_bookings": latest_bookings,
    }

@app.route("/admin")
def admin():
    auto_cancel_expired_bookings()

    if session.get("role") != "admin":
        return "สำหรับผู้ดูแลระบบเท่านั้น"

    conn = get_db()
    cur = conn.cursor()

    # เปิดหน้า "จัดการการจอง" แล้วให้ badge ของเมนูนี้หาย
    cur.execute("SELECT COALESCE(MAX(id), 0) AS max_id FROM notifications WHERE username = 'admin'")
    latest_notification = cur.fetchone()
    latest_notification_id = latest_notification["max_id"] if latest_notification else 0

    cur.execute("""
        UPDATE users
        SET last_seen_admin_bookings = ?, last_seen_admin_notification_id = ?
        WHERE username = ?
    """, (db_now_str(), latest_notification_id, session["user"]))

    # ตัวกรองรายการจองสำหรับผู้ดูแลระบบ
    filter_status = request.args.get("status", "all").strip()
    filter_type = request.args.get("type", "all").strip()
    filter_payment_type = request.args.get("payment_type", "all").strip()
    filter_room = request.args.get("room_id", "all").strip()
    filter_month = request.args.get("month", "all").strip()
    filter_year = request.args.get("year", "all").strip()
    search_q = request.args.get("q", "").strip()

    where = []
    params = []

    if filter_status and filter_status != "all":
        where.append("bookings.status = ?")
        params.append(filter_status)

    if filter_type == "walkin":
        where.append("bookings.username LIKE 'walkin_%'")
    elif filter_type == "online":
        where.append("bookings.username NOT LIKE 'walkin_%'")

    if filter_payment_type == "full":
        where.append("bookings.paytype = ?")
        params.append("full")
    elif filter_payment_type == "deposit":
        where.append("bookings.paytype = ?")
        params.append("deposit")
    elif filter_payment_type != "all":
        filter_payment_type = "all"

    if filter_room and filter_room != "all":
        try:
            where.append("bookings.room_id = ?")
            params.append(int(filter_room))
        except ValueError:
            filter_room = "all"

    if filter_year and filter_year != "all":
        if filter_year.isdigit():
            where.append("strftime('%Y', bookings.checkin) = ?")
            params.append(filter_year)
        else:
            filter_year = "all"

    if filter_month and filter_month != "all":
        if filter_month.isdigit():
            where.append("strftime('%m', bookings.checkin) = ?")
            params.append(f"{int(filter_month):02d}")
        else:
            filter_month = "all"

    if search_q:
        like = f"%{search_q}%"
        where.append("""
            (
                bookings.username LIKE ? OR
                bookings.room_name LIKE ? OR
                users.first_name LIKE ? OR
                users.last_name LIKE ? OR
                users.nickname LIKE ? OR
                users.phone LIKE ? OR
                users.email LIKE ?
            )
        """)
        params.extend([like] * 7)

    where_sql = "WHERE " + " AND ".join(where) if where else ""

    cur.execute(f"""
        SELECT bookings.*, users.first_name, users.last_name, users.phone AS user_phone, users.email AS user_email, users.nickname AS user_nickname, users.gender AS user_gender
        FROM bookings
        LEFT JOIN users ON bookings.username = users.username
        {where_sql}
        ORDER BY bookings.id DESC
    """, params)
    bookings = cur.fetchall()

    cur.execute("SELECT * FROM room_blocks ORDER BY start_date ASC")
    blocks = cur.fetchall()

    cur.execute("SELECT DISTINCT strftime('%Y', checkin) AS year FROM bookings WHERE checkin IS NOT NULL AND checkin != '' ORDER BY year DESC")
    year_rows = [row["year"] for row in cur.fetchall() if row["year"]]
    current_year = str(date.today().year)
    if current_year not in year_rows:
        year_rows.insert(0, current_year)

    conn.commit()
    conn.close()

    booking_filters = {
        "status": filter_status,
        "type": filter_type,
        "payment_type": filter_payment_type,
        "room_id": filter_room,
        "month": filter_month,
        "year": filter_year,
        "q": search_q,
        "years": year_rows,
        "result_count": len(bookings),
    }

    return render_template("admin.html", bookings=bookings, blocks=blocks, rooms=get_rooms_with_status(), report=get_admin_report_summary(), booking_filters=booking_filters)



@app.route("/admin/rooms")
def admin_rooms():
    if session.get("role") != "admin":
        return "สำหรับผู้ดูแลระบบเท่านั้น"
    return render_template(
        "admin_rooms.html",
        rooms=get_rooms_with_status(include_deleted=True),
        amenity_options=ROOM_AMENITY_OPTIONS
    )

@app.route("/admin/rooms/add", methods=["POST"])
def admin_add_room():
    if session.get("role") != "admin":
        return "สำหรับผู้ดูแลระบบเท่านั้น"

    name = request.form.get("name", "").strip()
    detail = request.form.get("detail", "").strip()
    status_text = request.form.get("status_text", "").strip()
    maintenance_note = request.form.get("maintenance_note", "").strip()
    try:
        price = float(request.form.get("price", "0") or 0)
    except ValueError:
        price = 0

    if not name:
        flash("กรุณากรอกชื่อห้อง")
        return redirect("/admin/rooms")
    if price <= 0:
        flash("กรุณากรอกราคาห้องให้ถูกต้อง")
        return redirect("/admin/rooms")

    cover = save_uploaded_room_file(request.files.get("cover_file")) or request.form.get("cover", "").strip() or "room1_1.jpg"
    images = []
    for file in request.files.getlist("image_files"):
        saved = save_uploaded_room_file(file)
        if saved:
            images.append(saved)
    manual_images = [x.strip() for x in request.form.get("images", "").replace("\n", ",").split(",") if x.strip()]
    images.extend(manual_images)
    if cover not in images:
        images.insert(0, cover)

    amenities = build_room_amenities_from_form()
    is_open = 1 if request.form.get("is_open") == "1" else 0

    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO rooms(name, price, detail, cover, images, amenities, is_open, is_deleted, status_text, maintenance_note, created_at, updated_at)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        name, price, detail, cover, json.dumps(images, ensure_ascii=False), json.dumps(amenities, ensure_ascii=False),
        is_open, 0, status_text, maintenance_note, db_now_str(), db_now_str()
    ))
    new_room_id = cur.lastrowid
    cur.execute("INSERT OR IGNORE INTO room_settings(room_id, is_open) VALUES (?, ?)", (new_room_id, is_open))
    conn.commit()
    conn.close()
    flash(f"เพิ่มห้อง {name} เรียบร้อยแล้ว")
    return redirect("/admin/rooms")


@app.route("/admin/rooms/<int:room_id>/edit", methods=["POST"])
def admin_edit_room(room_id):
    if session.get("role") != "admin":
        return "สำหรับผู้ดูแลระบบเท่านั้น"

    name = request.form.get("name", "").strip()
    detail = request.form.get("detail", "").strip()
    status_text = request.form.get("status_text", "").strip()
    maintenance_note = request.form.get("maintenance_note", "").strip()
    try:
        price = float(request.form.get("price", "0") or 0)
    except ValueError:
        price = 0

    if not name or price <= 0:
        flash("กรุณากรอกชื่อห้องและราคาห้องให้ถูกต้อง")
        return redirect("/admin/rooms")

    old_room = get_room(room_id)
    if not old_room:
        flash("ไม่พบห้องพัก")
        return redirect("/admin/rooms")

    cover = save_uploaded_room_file(request.files.get("cover_file")) or request.form.get("cover", "").strip() or old_room.get("cover") or "room1_1.jpg"
    images = []
    for file in request.files.getlist("image_files"):
        saved = save_uploaded_room_file(file)
        if saved:
            images.append(saved)
    manual_images = [x.strip() for x in request.form.get("images", "").replace("\n", ",").split(",") if x.strip()]
    if manual_images:
        images.extend(manual_images)
    else:
        images.extend(old_room.get("images", []))
    if cover not in images:
        images.insert(0, cover)

    amenities = build_room_amenities_from_form()
    is_open = 1 if request.form.get("is_open") == "1" else 0

    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        UPDATE rooms
        SET name=?, price=?, detail=?, cover=?, images=?, amenities=?, is_open=?, status_text=?, maintenance_note=?, updated_at=?
        WHERE id=?
    """, (
        name, price, detail, cover, json.dumps(images, ensure_ascii=False), json.dumps(amenities, ensure_ascii=False),
        is_open, status_text, maintenance_note, db_now_str(), room_id
    ))
    cur.execute(
        "INSERT INTO room_settings(room_id, is_open) VALUES (?, ?) ON CONFLICT(room_id) DO UPDATE SET is_open = excluded.is_open",
        (room_id, is_open)
    )
    conn.commit()
    conn.close()
    flash(f"แก้ไขข้อมูลห้อง {name} เรียบร้อยแล้ว")
    return redirect("/admin/rooms")


@app.route("/admin/rooms/<int:room_id>/delete", methods=["POST"])
def admin_delete_room(room_id):
    if session.get("role") != "admin":
        return "สำหรับผู้ดูแลระบบเท่านั้น"

    room = get_room(room_id)
    if not room:
        flash("ไม่พบห้องพัก")
        return redirect("/admin/rooms")

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) AS cnt FROM bookings WHERE room_id = ?", (room_id,))
    has_bookings = (cur.fetchone()["cnt"] or 0) > 0
    if has_bookings:
        cur.execute("UPDATE rooms SET is_deleted = 1, is_open = 0, updated_at = ? WHERE id = ?", (db_now_str(), room_id))
        cur.execute(
            "INSERT INTO room_settings(room_id, is_open) VALUES (?, 0) ON CONFLICT(room_id) DO UPDATE SET is_open = 0",
            (room_id,)
        )
        flash(f"ซ่อนห้อง {room['name']} แล้ว เนื่องจากมีประวัติการจองเดิม")
    else:
        cur.execute("DELETE FROM rooms WHERE id = ?", (room_id,))
        cur.execute("DELETE FROM room_settings WHERE room_id = ?", (room_id,))
        flash(f"ลบห้อง {room['name']} เรียบร้อยแล้ว")
    conn.commit()
    conn.close()
    return redirect("/admin/rooms")


@app.route("/admin/toggle-room/<int:room_id>")
def admin_toggle_room(room_id):
    if session.get("role") != "admin":
        return "สำหรับผู้ดูแลระบบเท่านั้น"

    room = None
    for r in get_rooms_with_status():
        if r["id"] == room_id:
            room = r
            break

    if room is None:
        flash("ไม่พบห้องพัก")
        return redirect("/admin")

    new_is_open = 0 if room.get("is_open", True) else 1

    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE rooms SET is_open = ?, updated_at = ? WHERE id = ?", (new_is_open, db_now_str(), room_id))
    cur.execute(
        "INSERT INTO room_settings(room_id, is_open) VALUES (?, ?) ON CONFLICT(room_id) DO UPDATE SET is_open = excluded.is_open",
        (room_id, new_is_open)
    )
    conn.commit()
    conn.close()

    if new_is_open:
        flash(f"เปิดห้อง {room['name']} ให้จองได้แล้ว")
    else:
        flash(f"ปิดห้อง {room['name']} ชั่วคราวแล้ว")

    return redirect("/admin")


@app.route("/admin/calendar/update-room-day", methods=["POST"])
def admin_calendar_update_room_day():
    if session.get("role") != "admin":
        return "สำหรับผู้ดูแลระบบเท่านั้น"

    return_url = request.form.get("return_url") or "/admin/calendar"
    if not return_url.startswith("/admin/calendar"):
        return_url = "/admin/calendar"

    try:
        room_id = int(request.form.get("room_id", ""))
        target_date_text = request.form.get("date", "")
        target_day = datetime.strptime(target_date_text, "%Y-%m-%d").date()
    except ValueError:
        flash("ข้อมูลห้องหรือวันที่ไม่ถูกต้อง")
        return redirect(return_url)

    action = request.form.get("action")
    selected_status = request.form.get("status", "blocked").strip()
    note = request.form.get("note", "").strip()

    if action == "block":
        selected_status = "blocked"
    elif action == "unblock":
        selected_status = "free"
    elif action != "set_status":
        flash("ไม่พบคำสั่งที่ต้องการ")
        return redirect(return_url)

    if selected_status not in MANUAL_DAY_STATUSES:
        flash("สถานะที่เลือกไม่ถูกต้อง")
        return redirect(return_url)

    conn = get_db()
    cur = conn.cursor()

    if selected_status != "free":
        placeholders = ",".join("?" for _ in BOOKED_STATUSES)
        cur.execute(f"""
            SELECT * FROM bookings
            WHERE room_id = ? AND status IN ({placeholders})
        """, (room_id, *BOOKED_STATUSES))
        bookings = cur.fetchall()

        for booking in bookings:
            checkin = datetime.strptime(f"{booking['checkin']} {booking['checkin_time']}", "%Y-%m-%d %H:%M")
            checkout = datetime.strptime(f"{booking['checkout']} {booking['checkout_time']}", "%Y-%m-%d %H:%M")
            if block_overlaps_booking(checkin, checkout, target_day, target_day):
                conn.close()
                flash("เปลี่ยนไม่ได้ เพราะวันดังกล่าวมีรายการจองหรือมีผู้พักอยู่")
                return redirect(return_url)

    cur.execute("""
        SELECT * FROM room_blocks
        WHERE room_id = ? AND date(start_date) <= date(?) AND date(end_date) >= date(?)
    """, (room_id, target_date_text, target_date_text))
    matching_blocks = cur.fetchall()

    for block in matching_blocks:
        block_start = datetime.strptime(block["start_date"], "%Y-%m-%d").date()
        block_end = datetime.strptime(block["end_date"], "%Y-%m-%d").date()
        block_status = block["status"] if "status" in block.keys() and block["status"] else "blocked"
        cur.execute("DELETE FROM room_blocks WHERE id = ?", (block["id"],))

        if block_start < target_day:
            before_end = target_day - timedelta(days=1)
            cur.execute("""
                INSERT INTO room_blocks(room_id, start_date, end_date, note, created_at, status)
                VALUES(?,?,?,?,?,?)
            """, (room_id, block_start.strftime("%Y-%m-%d"), before_end.strftime("%Y-%m-%d"), block["note"], block["created_at"], block_status))

        if target_day < block_end:
            after_start = target_day + timedelta(days=1)
            cur.execute("""
                INSERT INTO room_blocks(room_id, start_date, end_date, note, created_at, status)
                VALUES(?,?,?,?,?,?)
            """, (room_id, after_start.strftime("%Y-%m-%d"), block_end.strftime("%Y-%m-%d"), block["note"], block["created_at"], block_status))

    if selected_status != "free":
        cur.execute("""
            INSERT INTO room_blocks(room_id, start_date, end_date, note, created_at, status)
            VALUES(?,?,?,?,?,?)
        """, (room_id, target_date_text, target_date_text, note, get_now().strftime("%Y-%m-%d %H:%M:%S"), selected_status))
        flash(f"ตั้งค่าวันนี้เป็น {MANUAL_DAY_STATUSES[selected_status]} แล้ว")
    else:
        flash("เปลี่ยนเฉพาะวันดังกล่าวกลับเป็นว่างแล้ว")

    conn.commit()
    conn.close()
    return redirect(return_url)


@app.route("/admin/calendar")
def admin_calendar():
    auto_cancel_expired_bookings()

    if session.get("role") != "admin":
        return "สำหรับผู้ดูแลระบบเท่านั้น"

    start_text = request.args.get("start", "").strip()
    try:
        start_day = datetime.strptime(start_text, "%Y-%m-%d").date() if start_text else date.today()
    except ValueError:
        start_day = date.today()

    # แสดงเป็นช่วง 14 วันเพื่อให้ตารางไม่ล้นจอ แต่มีปุ่มเลื่อนไปเดือนหน้า/ช่วงถัดไปได้
    days = 14
    day_list, room_rows = get_calendar_data(days=days, start_day=start_day)
    prev_start = (start_day - timedelta(days=14)).strftime("%Y-%m-%d")
    next_start = (start_day + timedelta(days=14)).strftime("%Y-%m-%d")
    today_start = date.today().strftime("%Y-%m-%d")
    current_start = start_day.strftime("%Y-%m-%d")
    current_end = day_list[-1].strftime("%Y-%m-%d") if day_list else current_start
    return render_template(
        "admin_calendar.html",
        day_list=day_list,
        room_rows=room_rows,
        prev_start=prev_start,
        next_start=next_start,
        today_start=today_start,
        current_start=current_start,
        current_end=current_end,
    )


def _month_range(year, month):
    start = date(int(year), int(month), 1)
    if int(month) == 12:
        end = date(int(year) + 1, 1, 1)
    else:
        end = date(int(year), int(month) + 1, 1)
    return start, end


def _admin_monthly_report(year, month):
    """คำนวณรายงานย้อนหลังตามเดือน/ปีจากตาราง bookings

    หลักคิดด้านรายได้:
    - ยอดจองรวม = มูลค่ารวมของรายการที่ยังเกี่ยวข้องกับการให้บริการ
    - รายได้ที่รับแล้ว = เงินที่ได้รับการยืนยันแล้วเท่านั้น
      เช่น มัดจำ 25% จะนับเฉพาะยอดมัดจำก่อน จนกว่าจะรับยอดคงเหลือครบ
    - ยอดค้างชำระ = เงินที่ยังต้องเก็บเพิ่มจากลูกค้า
    """
    start_date, end_date = _month_range(year, month)
    start_text = start_date.strftime("%Y-%m-%d")
    end_text = end_date.strftime("%Y-%m-%d")

    # สถานะที่ยังถือว่าเกี่ยวข้องกับรายงาน ไม่รวมรายการที่ถูกยกเลิก/ปฏิเสธ/คืนเงินแล้ว
    active_statuses = (
        "approved", "waiting_slip_approval", "deposit_paid",
        "waiting_remaining_slip_approval", "remaining_counter_pending",
        "counter_pending", "fully_paid", "checked_in", "checked_out"
    )

    # สถานะที่ยืนยันว่ารับเงินบางส่วนแล้ว เช่น มัดจำ 25%
    partial_paid_statuses = (
        "deposit_paid", "waiting_remaining_slip_approval", "remaining_counter_pending"
    )

    # สถานะที่ยืนยันว่าชำระครบแล้ว
    full_paid_statuses = ("fully_paid", "checked_in", "checked_out")

    conn = get_db()
    cur = conn.cursor()

    placeholders_active = ",".join("?" for _ in active_statuses)
    placeholders_partial = ",".join("?" for _ in partial_paid_statuses)
    placeholders_full = ",".join("?" for _ in full_paid_statuses)

    # ยอดจองรวม: ใช้ total ของรายการที่ยังเกี่ยวข้องกับการให้บริการ
    cur.execute(f"""
        SELECT COALESCE(SUM(total), 0) AS total
        FROM bookings
        WHERE status IN ({placeholders_active})
          AND checkin >= ? AND checkin < ?
    """, (*active_statuses, start_text, end_text))
    booking_total = float((cur.fetchone() or {"total": 0})["total"] or 0)

    # รายได้ที่รับแล้ว: นับเฉพาะเงินที่ยืนยันแล้ว
    # - รอชำระ/รอตรวจสลิป/รอชำระหน้าเคาน์เตอร์ = ยังไม่นับรายได้
    # - มัดจำแล้ว = นับเฉพาะ paid
    # - ชำระครบ/เข้าพัก/เช็กเอาท์ = นับ paid ซึ่งระบบอัปเดตเป็นยอดเต็มแล้ว
    cur.execute(f"""
        SELECT COALESCE(SUM(
            CASE
                WHEN status IN ({placeholders_partial}) THEN COALESCE(paid, 0)
                WHEN status IN ({placeholders_full}) THEN COALESCE(paid, total, 0)
                ELSE 0
            END
        ), 0) AS total
        FROM bookings
        WHERE status IN ({placeholders_active})
          AND checkin >= ? AND checkin < ?
    """, (*partial_paid_statuses, *full_paid_statuses, *active_statuses, start_text, end_text))
    revenue_received = float((cur.fetchone() or {"total": 0})["total"] or 0)

    # ยอดค้างชำระ: รายการที่ยังไม่ได้รับเงินเลยให้นับเต็มจำนวน / รายการมัดจำให้นับ remaining_due
    cur.execute(f"""
        SELECT COALESCE(SUM(
            CASE
                WHEN status IN ('approved', 'waiting_slip_approval', 'counter_pending') THEN COALESCE(total, 0)
                WHEN status IN ({placeholders_partial}) THEN COALESCE(remaining_due, 0)
                ELSE 0
            END
        ), 0) AS total
        FROM bookings
        WHERE status IN ({placeholders_active})
          AND checkin >= ? AND checkin < ?
    """, (*partial_paid_statuses, *active_statuses, start_text, end_text))
    outstanding_revenue = float((cur.fetchone() or {"total": 0})["total"] or 0)

    cur.execute(f"""
        SELECT COUNT(*) AS total
        FROM bookings
        WHERE status IN ({placeholders_active})
          AND checkin >= ? AND checkin < ?
    """, (*active_statuses, start_text, end_text))
    booking_count = int((cur.fetchone() or {"total": 0})["total"] or 0)

    cur.execute("""
        SELECT COUNT(*) AS total
        FROM bookings
        WHERE username LIKE 'walkin_%'
          AND checkin >= ? AND checkin < ?
    """, (start_text, end_text))
    walkin_count = int((cur.fetchone() or {"total": 0})["total"] or 0)

    cur.execute(f"""
        SELECT COUNT(DISTINCT username) AS total
        FROM bookings
        WHERE status IN ({placeholders_active})
          AND username NOT LIKE 'walkin_%'
          AND checkin >= ? AND checkin < ?
    """, (*active_statuses, start_text, end_text))
    customer_count = int((cur.fetchone() or {"total": 0})["total"] or 0)

    cur.execute(f"""
        SELECT COALESCE(SUM(nights), 0) AS total
        FROM bookings
        WHERE status IN ({placeholders_active})
          AND checkin >= ? AND checkin < ?
    """, (*active_statuses, start_text, end_text))
    total_nights = int(float((cur.fetchone() or {"total": 0})["total"] or 0))

    cur.execute(f"""
        SELECT room_name, COUNT(*) AS total
        FROM bookings
        WHERE status IN ({placeholders_active})
          AND checkin >= ? AND checkin < ?
        GROUP BY room_name
        ORDER BY total DESC
        LIMIT 1
    """, (*active_statuses, start_text, end_text))
    top_room_row = cur.fetchone()
    top_room = top_room_row["room_name"] if top_room_row else "-"

    cur.execute("""
        SELECT COUNT(*) AS total
        FROM room_blocks
        WHERE NOT (end_date < ? OR start_date >= ?)
    """, (start_text, end_text))
    blocked_count = int((cur.fetchone() or {"total": 0})["total"] or 0)

    cur.execute("""
        SELECT room_name, username, checkin, checkout, total, paid, remaining_due, status, created_at
        FROM bookings
        WHERE checkin >= ? AND checkin < ?
        ORDER BY id DESC
        LIMIT 10
    """, (start_text, end_text))
    latest_bookings = cur.fetchall()

    # รายได้แยกรายวัน: แสดงทั้งยอดจองรวม รายได้ที่รับแล้ว และยอดค้างชำระ
    cur.execute(f"""
        SELECT
            checkin AS day,
            COUNT(*) AS count,
            COALESCE(SUM(total), 0) AS booking_total,
            COALESCE(SUM(
                CASE
                    WHEN status IN ({placeholders_partial}) THEN COALESCE(paid, 0)
                    WHEN status IN ({placeholders_full}) THEN COALESCE(paid, total, 0)
                    ELSE 0
                END
            ), 0) AS revenue_received,
            COALESCE(SUM(
                CASE
                    WHEN status IN ('approved', 'waiting_slip_approval', 'counter_pending') THEN COALESCE(total, 0)
                    WHEN status IN ({placeholders_partial}) THEN COALESCE(remaining_due, 0)
                    ELSE 0
                END
            ), 0) AS outstanding_revenue
        FROM bookings
        WHERE status IN ({placeholders_active})
          AND checkin >= ? AND checkin < ?
        GROUP BY checkin
        ORDER BY checkin ASC
    """, (*partial_paid_statuses, *full_paid_statuses, *partial_paid_statuses, *active_statuses, start_text, end_text))
    daily_rows = cur.fetchall()

    conn.close()

    month_names = [
        "", "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน",
        "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"
    ]

    return {
        "year": int(year),
        "month": int(month),
        "month_name": month_names[int(month)],
        "period_text": f"{month_names[int(month)]} {int(year)}",
        "revenue": revenue_received,
        "revenue_received": revenue_received,
        "outstanding_revenue": outstanding_revenue,
        "booking_total": booking_total,
        "booking_count": booking_count,
        "walkin_count": walkin_count,
        "customer_count": customer_count,
        "total_nights": total_nights,
        "top_room": top_room,
        "blocked_count": blocked_count,
        "latest_bookings": latest_bookings,
        "daily_rows": daily_rows,
    }


@app.route("/admin/reports")
def admin_reports():
    if session.get("role") != "admin":
        flash("กรุณาเข้าสู่ระบบผู้ดูแลระบบก่อน")
        return redirect("/login")

    today = date.today()
    current_year = today.year
    current_month = today.month

    try:
        selected_month = int(request.args.get("month", current_month))
        selected_year = int(request.args.get("year", current_year))
    except ValueError:
        selected_month = current_month
        selected_year = current_year

    # รายงานย้อนหลัง: อนุญาตให้เลือกได้เฉพาะเดือนปัจจุบันและเดือนเก่าเท่านั้น
    # กันไม่ให้เลือกเดือน/ปีอนาคตจากทั้ง Dropdown และการแก้ URL เอง
    min_year = current_year - 5
    if selected_year > current_year:
        selected_year = current_year
    if selected_year < min_year:
        selected_year = min_year
    if selected_month < 1 or selected_month > 12:
        selected_month = current_month
    if selected_year == current_year and selected_month > current_month:
        selected_month = current_month

    available_months = list(range(1, current_month + 1)) if selected_year == current_year else list(range(1, 13))
    years = list(range(current_year, min_year - 1, -1))

    report = _admin_monthly_report(selected_year, selected_month)
    return render_template(
        "admin_reports.html",
        report=report,
        selected_month=selected_month,
        selected_year=selected_year,
        current_month=current_month,
        current_year=current_year,
        available_months=available_months,
        years=years,
    )


init_db()



if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)